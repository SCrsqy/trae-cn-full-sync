import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import os

# 创建工作簿
wb = Workbook()

# 移除默认工作表
default_sheet = wb.active
wb.remove(default_sheet)

# 创建项目封面工作表
sheet1 = wb.create_sheet(title="项目封面")
sheet1['A4'] = "中天建设集团有限公司"
sheet1['A5'] = "成华区迎晖路194号住宅、商业及配套设施项目"
sheet1['A6'] = "幕墙专业分包工程"
sheet1['A7'] = "项目成本管理策划书"
sheet1['A8'] = "编制部门： 项目部/成本合约部"
sheet1['A9'] = "编制日期： 2024年X月X日"
sheet1['A10'] = "版本号： V1.0"
sheet1['A11'] = "【内部机密 请勿外传】"

# 设置字体和对齐
for row in range(4, 12):
    cell = sheet1[f'A{row}']
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')

# 创建项目概况工作表
sheet2 = wb.create_sheet(title="项目概况")
sheet2['A15'] = "一、项目基本信息"
sheet2['A16'] = "项目名称"
sheet2['B16'] = "成华区迎晖路194号住宅、商业及配套设施项目幕墙工程"
sheet2['A17'] = "工程地点"
sheet2['B17'] = "成都市成华区迎晖路194号"
sheet2['A18'] = "建设单位"
sheet2['B18'] = "长住集团"
sheet2['A19'] = "总包单位"
sheet2['B19'] = "中天建设集团有限公司西南公司"
sheet2['A20'] = "二、工程范围与规模"
sheet2['A21'] = "工程范围"
sheet2['B21'] = "本项目包含排屋（11#楼）、1#2#商业及1#2#塔楼的幕墙工程。"
sheet2['A22'] = "主要幕墙形式"
sheet2['B22'] = "玻璃幕墙（座槽式、T型钢式、明框式）、石材幕墙、铝板幕墙、铝合金门窗、格栅、玻璃栏杆、采光顶等。"
sheet2['A23'] = "暂估幕墙面积"
sheet2['B23'] = "1. 排屋: 待定 (详见明细)"
sheet2['B24'] = "2. 1#2#商业: 待定 (详见明细)"
sheet2['B25'] = "3. 1#2#塔楼: 30,816.26 ㎡"
sheet2['A26'] = "三、合同与造价"
sheet2['A27'] = "合同形式"
sheet2['B27'] = "暂定工程量，固定综合单价"
sheet2['A28'] = "项目总预算"
sheet2['B28'] = "84,002,317 元 (暂估，含塔楼)"
sheet2['A29'] = "四、项目重难点"
sheet2['A30'] = "1. 材料复杂性"
sheet2['B30'] = "石材、铝板、玻璃种类繁多，且存在大量弧形、异形板，加工和安装难度高。"
sheet2['A31'] = "2. 图纸不完善"
sheet2['B31'] = "目前图纸不齐全，工程量存在较大不确定性，是主要风险源。"
sheet2['A32'] = "3. 工期紧张"
sheet2['B32'] = "展示区要求高，存在赶工和夜间施工情况，将增加措施及管理成本。"
sheet2['A33'] = "4. 质量控制"
sheet2['B33'] = "高端住宅项目，对幕墙的观感质量、防水性能等要求极高。"

# 设置字体和对齐
for row in range(15, 34):
    if row in [15, 20, 26, 29]:
        cell = sheet2[f'A{row}']
        cell.font = Font(size=12, bold=True)
    else:
        cell = sheet2[f'A{row}']
        cell.font = Font(size=11)
    sheet2[f'B{row}'].alignment = Alignment(wrap_text=True)

# 创建成本利润基础数据工作表
sheet3 = wb.create_sheet(title="成本利润基础数据")
sheet3['A37'] = "一、成本汇总（不含税）"
sheet3['A38'] = "序号"
sheet3['B38'] = "项目"
sheet3['C38'] = "金额（元）"
sheet3['D38'] = "占总成本比例"
sheet3['E38'] = "备注"
sheet3['A39'] = "1"
sheet3['B39'] = "排屋幕墙直接成本"
sheet3['C39'] = 18880920
sheet3['D39'] = "23%"
sheet3['E39'] = "基于工程量清单反算"
sheet3['A40'] = "2"
sheet3['B40'] = "商业幕墙直接成本"
sheet3['C40'] = 25160812
sheet3['D40'] = "30%"
sheet3['E40'] = "基于工程量清单反算"
sheet3['A41'] = "3"
sheet3['B41'] = "塔楼幕墙直接成本"
sheet3['C41'] = 25390235
sheet3['D41'] = "31%"
sheet3['E41'] = "基于面积及单方造价估算"
sheet3['A42'] = "4"
sheet3['B42'] = "措施费"
sheet3['C42'] = 921500
sheet3['D42'] = "1%"
sheet3['E42'] = "脚手架、吊篮、搬运等"
sheet3['A43'] = "5"
sheet3['B43'] = "管理费"
sheet3['C43'] = 4900000
sheet3['D43'] = "6%"
sheet3['E43'] = "管理人员工资、差旅、办公等"
sheet3['A44'] = "6"
sheet3['B44'] = "规费及税金"
sheet3['C44'] = 7500000
sheet3['D44'] = "9%"
sheet3['E44'] = "按9%税率估算"
sheet3['A45'] = "7"
sheet3['B45'] = "项目总成本"
sheet3['C45'] = 82753467
sheet3['D45'] = "100%"
sheet3['A47'] = "二、收入与利润测算"
sheet3['A48'] = "1"
sheet3['B48'] = "预计总收入"
sheet3['C48'] = 84002317
sheet3['E48'] = "参考中标价或与业主暂定价"
sheet3['A49'] = "2"
sheet3['B49'] = "预计总成本"
sheet3['C49'] = 82753467
sheet3['A50'] = "3"
sheet3['B50'] = "预计利润"
sheet3['C50'] = 1248850
sheet3['A51'] = "4"
sheet3['B51'] = "预计利润率"
sheet3['C51'] = "1.49%"
sheet3['E51'] = "利润空间极小，成本控制是关键！"

# 设置字体和对齐
for row in range(37, 52):
    if row in [37, 47]:
        cell = sheet3[f'A{row}']
        cell.font = Font(size=12, bold=True)
    elif row == 38:
        for col in range(1, 6):
            cell = sheet3[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        cell = sheet3[f'A{row}']
        cell.font = Font(size=11)

# 创建成本明细分析工作表
sheet4 = wb.create_sheet(title="成本明细分析")
sheet4['A54'] = "序号"
sheet4['B54'] = "区域"
sheet4['C54'] = "项目名称"
sheet4['D54'] = "单位"
sheet4['E54'] = "工程量"
sheet4['F54'] = "预算收入"
sheet4['G54'] = ""
sheet4['H54'] = "目标成本"
sheet4['I54'] = ""
sheet4['J54'] = ""
sheet4['K54'] = ""
sheet4['L54'] = ""
sheet4['M54'] = "盈亏分析"
sheet4['F55'] = "单价"
sheet4['G55'] = "合价"
sheet4['H55'] = "其中:主材费"
sheet4['I55'] = "人工费"
sheet4['J55'] = "辅材机械费"
sheet4['K55'] = "管理费"
sheet4['L55'] = "合价"
sheet4['M55'] = "盈亏"
sheet4['M56'] = "备注"
sheet4['A56'] = "1"
sheet4['B56'] = "排屋"
sheet4['C56'] = "80系列断桥隔热铝合金窗"
sheet4['D56'] = "m²"
sheet4['E56'] = 139
sheet4['F56'] = 950
sheet4['G56'] = 131602
sheet4['H56'] = 570
sheet4['I56'] = 152
sheet4['J56'] = 38
sheet4['K56'] = 9
sheet4['L56'] = 106483
sheet4['M56'] = 25119
sheet4['A57'] = "2"
sheet4['B57'] = "排屋"
sheet4['C57'] = "幕墙式铝合金窗"
sheet4['D57'] = "m²"
sheet4['E57'] = 1304
sheet4['F57'] = 1350
sheet4['G57'] = 1760097
sheet4['H57'] = 810
sheet4['I57'] = 216
sheet4['J57'] = 54
sheet4['K57'] = 13
sheet4['L57'] = 1425472
sheet4['M57'] = 334625
sheet4['A58'] = "3"
sheet4['B58'] = "排屋"
sheet4['C58'] = "石材幕墙"
sheet4['D58'] = "m²"
sheet4['E58'] = 1202
sheet4['F58'] = 1250
sheet4['G58'] = 1502375
sheet4['H58'] = 750
sheet4['I58'] = 200
sheet4['J58'] = 50
sheet4['K58'] = 12
sheet4['L58'] = 1216424
sheet4['M58'] = 285951
sheet4['A59'] = "4"
sheet4['B59'] = "排屋"
sheet4['C59'] = "铝板幕墙"
sheet4['D59'] = "m²"
sheet4['E59'] = 5411
sheet4['F59'] = 750
sheet4['G59'] = 4058115
sheet4['H59'] = 450
sheet4['I59'] = 120
sheet4['J59'] = 30
sheet4['K59'] = 7
sheet4['L59'] = 3284787
sheet4['M59'] = 773328
sheet4['A60'] = "5"
sheet4['B60'] = "排屋"
sheet4['C60'] = "铝合金格栅/百叶"
sheet4['D60'] = "m²"
sheet4['E60'] = 199
sheet4['F60'] = 650
sheet4['G60'] = 129157
sheet4['H60'] = 390
sheet4['I60'] = 104
sheet4['J60'] = 26
sheet4['K60'] = 6
sheet4['L60'] = 104544
sheet4['M60'] = 24613
sheet4['A61'] = "..."
sheet4['B61'] = "..."
sheet4['C61'] = "排屋小计"
sheet4['G61'] = 21532248
sheet4['L61'] = 17225798
sheet4['M61'] = 4306450
sheet4['A62'] = "10"
sheet4['B62'] = "商业"
sheet4['C62'] = "玻璃幕墙"
sheet4['D62'] = "m²"
sheet4['E62'] = 7701
sheet4['F62'] = 1900
sheet4['G62'] = 14631145
sheet4['H62'] = 1140
sheet4['I62'] = 304
sheet4['J62'] = 76
sheet4['K62'] = 19
sheet4['L62'] = 11844718
sheet4['M62'] = 2786427
sheet4['A63'] = "11"
sheet4['B63'] = "商业"
sheet4['C63'] = "石材幕墙"
sheet4['D63'] = "m²"
sheet4['E63'] = 1000
sheet4['F63'] = 1350
sheet4['G63'] = 1349652
sheet4['H63'] = 810
sheet4['I63'] = 216
sheet4['J63'] = 54
sheet4['K63'] = 13
sheet4['L63'] = 1092750
sheet4['M63'] = 256902
sheet4['A64'] = "12"
sheet4['B64'] = "商业"
sheet4['C64'] = "金属幕墙（铝板/网）"
sheet4['D64'] = "m²"
sheet4['E64'] = 10632
sheet4['F64'] = 780
sheet4['G64'] = 8293166
sheet4['H64'] = 468
sheet4['I64'] = 124
sheet4['J64'] = 31
sheet4['K64'] = 7
sheet4['L64'] = 6711840
sheet4['M64'] = 1581326
sheet4['A65'] = "..."
sheet4['B65'] = "..."
sheet4['C65'] = "商业小计"
sheet4['G65'] = 28683326
sheet4['L65'] = 23233494
sheet4['M65'] = 5449832
sheet4['A66'] = "20"
sheet4['B66'] = "塔楼"
sheet4['C66'] = "塔楼幕墙"
sheet4['D66'] = "m²"
sheet4['E66'] = 30816
sheet4['F66'] = 1018
sheet4['G66'] = 31370688
sheet4['L66'] = 25096550
sheet4['M66'] = 6274138
sheet4['M66'].comment = openpyxl.comments.Comment("单价按商业综合单价*80%估算", "User")
sheet4['A67'] = "直接费合计"
sheet4['G67'] = 81586262
sheet4['L67'] = 65555842
sheet4['M67'] = 16030420

# 设置字体和对齐
for row in range(54, 68):
    if row in [54, 55]:
        for col in range(1, 14):
            cell = sheet4[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 14):
            cell = sheet4[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 创建利润与风险分析工作表
sheet5 = wb.create_sheet(title="利润与风险分析")
sheet5['A71'] = "一、利润敏感性分析"
sheet5['A72'] = "变动因素"
sheet5['B72'] = "变动幅度"
sheet5['C72'] = "对总成本影响（元）"
sheet5['D72'] = "利润率影响"
sheet5['E72'] = "备注"
sheet5['A73'] = "主材价格上涨"
sheet5['B73'] = "+5%"
sheet5['C73'] = "+2,500,000"
sheet5['D73'] = "-2.98%"
sheet5['E73'] = "风险极高，铝材、玻璃价格波动大"
sheet5['A74'] = "人工成本上涨"
sheet5['B74'] = "+10%"
sheet5['C74'] = "+800,000"
sheet5['D74'] = "-0.95%"
sheet5['E74'] = "赶工期导致人工成本增加"
sheet5['A75'] = "工程量增加"
sheet5['B75'] = "+5%"
sheet5['C75'] = "+3,100,000"
sheet5['D75'] = "-3.69%"
sheet5['E75'] = "图纸不齐，变更风险巨大"
sheet5['A76'] = "材料损耗率控制"
sheet5['B76'] = "-1%"
sheet5['C76'] = "-500,000"
sheet5['D76'] = "+0.60%"
sheet5['E76'] = "通过精细化管理，可显著降本增效"
sheet5['A77'] = "二、主要风险识别与应对"
sheet5['A78'] = "风险类别"
sheet5['B78'] = "风险描述"
sheet5['C78'] = "可能性"
sheet5['D78'] = "影响程度"
sheet5['E78'] = "应对措施"
sheet5['A79'] = "工程量风险"
sheet5['B79'] = "图纸不完善，暂定工程量与实际差异大"
sheet5['C79'] = "高"
sheet5['D79'] = "高"
sheet5['E79'] = "1. 尽快与业主/设计确认图纸和工程量。\n2. 合同中明确变更索赔条款。\n3. 内部进行精细化翻样，提前预警。"
sheet5['A80'] = "材料风险"
sheet5['B80'] = "石材、铝板、弯弧玻璃等加工周期长、损耗大、价格波动"
sheet5['C80'] = "高"
sheet5['D80'] = "高"
sheet5['E80'] = "1. 提前锁定大宗材料价格。\n2. 优化排版，减少损耗。\n3. 派驻专人驻场监造，保证质量和进度。"
sheet5['A81'] = "工期风险"
sheet5['B81'] = "展示区工期紧张，需赶工"
sheet5['C81'] = "高"
sheet5['D81'] = "中"
sheet5['E81'] = "1. 制定详细可行的进度计划。\n2. 储备应急班组，应对高峰期。\n3. 赶工方案提前报批，保留索赔权利。"
sheet5['A82'] = "质量风险"
sheet5['B82'] = "弧形、异形板块安装精度难控制"
sheet5['C82'] = "中"
sheet5['D82'] = "高"
sheet5['E82'] = "1. 采用高精度测量放线仪器。\n2. 编制专项施工方案，样板引路。\n3. 加强过程验收。"

# 设置字体和对齐
for row in range(71, 83):
    if row in [71, 77]:
        cell = sheet5[f'A{row}']
        cell.font = Font(size=12, bold=True)
    elif row in [72, 78]:
        for col in range(1, 6):
            cell = sheet5[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 6):
            cell = sheet5[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)
    sheet5[f'E{row}'].alignment = Alignment(wrap_text=True)

# 创建分包谈判建议工作表
sheet6 = wb.create_sheet(title="分包谈判建议")
sheet6['A94'] = "谈判项"
sheet6['B94'] = "我方目标"
sheet6['C94'] = "分包商心理价位"
sheet6['D94'] = "谈判策略"
sheet6['E94'] = "负责人"
sheet6['A95'] = "劳务安装单价"
sheet6['B95'] = "在现有预算基础上下浮 5%-10%"
sheet6['C95'] = "预计可接受下浮 3%-5%"
sheet6['D95'] = "1. 以量换价：强调本项目体量大，系统多，可保证长期稳定用工。\n2. 难度加价反制：承认弧形、异形板块安装难度，但指出我方将提供完善的技术支持和测量放线，降低其管理成本。\n3. 付款条件：承诺进度款支付及时，换取价格优惠。"
sheet6['E95'] = "项目经理、成本经理"
sheet6['A96'] = "主材采购"
sheet6['B96'] = "争取 10%-15% 的下浮空间"
sheet6['C96'] = "不同材料下浮空间不同，石材5%，铝材8%，玻璃10%"
sheet6['D96'] = "1. 集中采购：整合排屋、商业、塔楼所有相同材料需求，统一招标。\n2. 引入竞争：每个品类至少邀请3家以上合格供应商报价。\n3. 战略合作：承诺后续项目优先合作权。"
sheet6['E96'] = "采购经理"
sheet6['A97'] = "措施费"
sheet6['B97'] = "打包总价，或按实结算部分争取 10% 下浮"
sheet6['C97'] = "5%左右"
sheet6['D97'] = "1. 对脚手架、吊篮等大型设备，利用集团长期合作商优势谈价。\n2. 对零星措施项，尽量包含在劳务单价中，减少单独计价。"
sheet6['E97'] = "生产经理、成本经理"
sheet6['A98'] = "付款条件"
sheet6['B98'] = "月进度80%，结算95%"
sheet6['C98'] = "月进度85%，结算95%"
sheet6['D98'] = "1. 以公司良好的付款信誉作为谈判筹码。\n2. 对于配合度高的分包商，可适当放宽付款比例。"
sheet6['E98'] = "项目经理、财务"

# 设置字体和对齐
for row in range(94, 99):
    if row == 94:
        for col in range(1, 6):
            cell = sheet6[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 6):
            cell = sheet6[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)
    sheet6[f'D{row}'].alignment = Alignment(wrap_text=True)

# 创建专项预算工作表
# 刚挂石材专项预算
sheet8 = wb.create_sheet(title="刚挂石材专项预算")
sheet8['A133'] = "序号"
sheet8['B133'] = "项目名称"
sheet8['C133'] = "单位"
sheet8['D133'] = "工程量"
sheet8['E133'] = "主材费"
sheet8['F133'] = "劳务费"
sheet8['G133'] = "综合单价"
sheet8['A134'] = "1"
sheet8['B134'] = "25mm葡萄牙灰石材"
sheet8['C134'] = "m²"
sheet8['D134'] = 500
sheet8['E134'] = 680
sheet8['F134'] = 210
sheet8['G134'] = 890
sheet8['A135'] = "2"
sheet8['B135'] = "25mm葡萄牙米黄石材"
sheet8['C135'] = "m²"
sheet8['D135'] = 800
sheet8['E135'] = 380
sheet8['F135'] = 210
sheet8['G135'] = 590
sheet8['A136'] = "..."
sheet8['B136'] = "..."
sheet8['C136'] = "..."
sheet8['D136'] = "..."
sheet8['E136'] = "..."
sheet8['F136'] = "..."
sheet8['G136'] = "..."

# 玻璃幕墙专项预算
sheet9 = wb.create_sheet(title="玻璃幕墙专项预算")
sheet9['A140'] = "序号"
sheet9['B140'] = "项目名称"
sheet9['C140'] = "单位"
sheet9['D140'] = "工程量"
sheet9['E140'] = "主材费"
sheet9['F140'] = "劳务费"
sheet9['G140'] = "综合单价"
sheet9['A141'] = "1"
sheet9['B141'] = "座槽式幕墙 (12+1.9SGP+12夹胶)"
sheet9['C141'] = "m²"
sheet9['D141'] = 500
sheet9['E141'] = 1200
sheet9['F141'] = 190
sheet9['G141'] = 1390
sheet9['A142'] = "2"
sheet9['B142'] = "T型钢幕墙 (8+1.52PVB+8夹胶)"
sheet9['C142'] = "m²"
sheet9['D142'] = 800
sheet9['E142'] = 700
sheet9['F142'] = 180
sheet9['G142'] = 880
sheet9['A143'] = "..."
sheet9['B143'] = "..."
sheet9['C143'] = "..."
sheet9['D143'] = "..."
sheet9['E143'] = "..."
sheet9['F143'] = "..."
sheet9['G143'] = "..."

# 其他专项安装预算
sheet10 = wb.create_sheet(title="其他专项安装预算")
sheet10['A147'] = "序号"
sheet10['B147'] = "项目名称"
sheet10['C147'] = "单位"
sheet10['D147'] = "工程量"
sheet10['E147'] = "主材费"
sheet10['F147'] = "劳务费"
sheet10['G147'] = "综合单价"
sheet10['A148'] = "1"
sheet10['B148'] = "铝合金格栅"
sheet10['C148'] = "m²"
sheet10['D148'] = 200
sheet10['E148'] = 450
sheet10['F148'] = 150
sheet10['G148'] = 600
sheet10['A149'] = "2"
sheet10['B149'] = "玻璃栏杆"
sheet10['C149'] = "m"
sheet10['D149'] = 500
sheet10['E149'] = 400
sheet10['F149'] = 135
sheet10['G149'] = 535
sheet10['A150'] = "..."
sheet10['B150'] = "..."
sheet10['C150'] = "..."
sheet10['D150'] = "..."
sheet10['E150'] = "..."
sheet10['F150'] = "..."
sheet10['G150'] = "..."

# 设置专项预算工作表的字体和对齐
for sheet in [sheet8, sheet9, sheet10]:
    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row):
        for cell in row:
            if cell.row == sheet.min_row:
                cell.font = Font(size=11, bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            else:
                cell.font = Font(size=11)

# 创建材料清单明细工作表
sheet11 = wb.create_sheet(title="材料清单明细")
sheet11['A156'] = "序号"
sheet11['B156'] = "类别"
sheet11['C156'] = "材料名称"
sheet11['D156'] = "规格型号"
sheet11['E156'] = "单位"
sheet11['F156'] = "数量"
sheet11['G156'] = "单价(含税)"
sheet11['H156'] = "合价"
sheet11['A157'] = "1"
sheet11['B157'] = "石材"
sheet11['C157'] = "25mm葡萄牙灰石材"
sheet11['E157'] = "m²"
sheet11['F157'] = 120
sheet11['G157'] = 768
sheet11['H157'] = 92203
sheet11['A158'] = "2"
sheet11['B158'] = "石材"
sheet11['C158'] = "35mm蓝眼睛石材（凹槽）"
sheet11['E158'] = "m²"
sheet11['F158'] = 50
sheet11['G158'] = 655
sheet11['H158'] = 32770
sheet11['A159'] = "..."
sheet11['B159'] = "..."
sheet11['C159'] = "..."
sheet11['D159'] = "..."
sheet11['E159'] = "..."
sheet11['F159'] = "..."
sheet11['G159'] = "..."
sheet11['H159'] = "..."
sheet11['A160'] = "100"
sheet11['B160'] = "铝型材"
sheet11['C160'] = "粉末喷涂铝型材T5"
sheet11['E160'] = "kg"
sheet11['F160'] = 20000
sheet11['G160'] = 27
sheet11['H160'] = 540000
sheet11['A161'] = "101"
sheet11['B161'] = "铝型材"
sheet11['C161'] = "阳极氧化铝型材T6"
sheet11['E161'] = "kg"
sheet11['F161'] = 5000
sheet11['G161'] = 30
sheet11['H161'] = 150000
sheet11['A162'] = "..."
sheet11['B162'] = "..."
sheet11['C162'] = "..."
sheet11['D162'] = "..."
sheet11['E162'] = "..."
sheet11['F162'] = "..."
sheet11['G162'] = "..."
sheet11['H162'] = "..."
sheet11['A163'] = "200"
sheet11['B163'] = "玻璃"
sheet11['C163'] = "12Low-E+12A+12 中空"
sheet11['E163'] = "m²"
sheet11['F163'] = 800
sheet11['G163'] = 299
sheet11['H163'] = 239200
sheet11['A164'] = "201"
sheet11['B164'] = "玻璃"
sheet11['C164'] = "15+2.28SGP+15 夹胶（超宽）"
sheet11['E164'] = "m²"
sheet11['F164'] = 100
sheet11['G164'] = 1695
sheet11['H164'] = 169500
sheet11['A165'] = "..."
sheet11['B165'] = "..."
sheet11['C165'] = "..."
sheet11['D165'] = "..."
sheet11['E165'] = "..."
sheet11['F165'] = "..."
sheet11['G165'] = "..."
sheet11['H165'] = "..."

# 设置字体和对齐
for row in range(156, 166):
    if row == 156:
        for col in range(1, 9):
            cell = sheet11[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 9):
            cell = sheet11[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 创建分包下浮比例分析工作表
sheet12 = wb.create_sheet(title="分包下浮比例分析")
sheet12['A169'] = "分包项目"
sheet12['B169'] = "我方预算总价"
sheet12['C169'] = "分包首次报价"
sheet12['D169'] = "谈判后目标价"
sheet12['E169'] = "目标下浮比例"
sheet12['F169'] = "责任人"
sheet12['A170'] = "排屋劳务"
sheet12['B170'] = 3200000
sheet12['C170'] = 3800000
sheet12['D170'] = 3500000
sheet12['E170'] = "7.8%"
sheet12['F170'] = "李经理"
sheet12['A171'] = "商业劳务"
sheet12['B171'] = 4500000
sheet12['C171'] = 5200000
sheet12['D171'] = 4800000
sheet12['E171'] = "7.7%"
sheet12['F171'] = "李经理"
sheet12['A172'] = "石材供应"
sheet12['B172'] = 2100000
sheet12['C172'] = 2300000
sheet12['D172'] = 2000000
sheet12['E172'] = "13.0%"
sheet12['F172'] = "王经理"
sheet12['A173'] = "铝材供应"
sheet12['B173'] = 3500000
sheet12['C173'] = 3800000
sheet12['D173'] = 3300000
sheet12['E173'] = "13.2%"
sheet12['F173'] = "王经理"
sheet12['A174'] = "..."
sheet12['B174'] = "..."
sheet12['C174'] = "..."
sheet12['D174'] = "..."
sheet12['E174'] = "..."
sheet12['F174'] = "..."

# 设置字体和对齐
for row in range(169, 175):
    if row == 169:
        for col in range(1, 7):
            cell = sheet12[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 7):
            cell = sheet12[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 创建分包分析工作表
sheet13 = wb.create_sheet(title="分包分析")
sheet13['A178'] = "分包商名称"
sheet13['B178'] = "承包范围"
sheet13['C178'] = "报价（元）"
sheet13['D178'] = "技术能力"
sheet13['E178'] = "配合度"
sheet13['F178'] = "综合评分"
sheet13['A179'] = "XX劳务公司"
sheet13['B179'] = "排屋劳务"
sheet13['C179'] = 3800000
sheet13['D179'] = "B+"
sheet13['E179'] = "A-"
sheet13['F179'] = 8.5
sheet13['A180'] = "YY劳务公司"
sheet13['B180'] = "商业劳务"
sheet13['C180'] = 5200000
sheet13['D180'] = "A-"
sheet13['E180'] = "B+"
sheet13['F180'] = 8.8
sheet13['A181'] = "ZZ石材公司"
sheet13['B181'] = "石材供应"
sheet13['C181'] = 2300000
sheet13['D181'] = "A"
sheet13['E181'] = "A"
sheet13['F181'] = 9.2
sheet13['A182'] = "AA铝材公司"
sheet13['B182'] = "铝材供应"
sheet13['C182'] = 3800000
sheet13['D182'] = "A-"
sheet13['E182'] = "A-"
sheet13['F182'] = 8.9
sheet13['A183'] = "..."
sheet13['B183'] = "..."
sheet13['C183'] = "..."
sheet13['D183'] = "..."
sheet13['E183'] = "..."
sheet13['F183'] = "..."
sheet13['A184'] = "结论与选择"
sheet13['A185'] = "经综合评估，建议选择："
sheet13['A186'] = "1. YY劳务公司作为商业劳务承包商，其技术能力更强，能更好应对复杂节点。"
sheet13['A187'] = "2. ZZ石材公司虽然价格不是最低，但其质量和配合度最佳，可有效降低损耗风险。"

# 设置字体和对齐
for row in range(178, 188):
    if row == 178:
        for col in range(1, 7):
            cell = sheet13[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    elif row in [184, 185]:
        cell = sheet13[f'A{row}']
        cell.font = Font(size=11, bold=True)
    else:
        for col in range(1, 7):
            cell = sheet13[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 创建动态成本分析表工作表
sheet14 = wb.create_sheet(title="中天建设塔子山幕墙项目动态成本分析表")
sheet14['A191'] = "成本科目"
sheet14['B191'] = "预算成本"
sheet14['C191'] = "已发生成本"
sheet14['D191'] = "待发生成本"
sheet14['E191'] = "动态总成本"
sheet14['F191'] = "偏差"
sheet14['G191'] = "偏差率"
sheet14['H191'] = "责任人"
sheet14['I191'] = "趋势"
sheet14['J191'] = "预警"
sheet14['A192'] = "排屋-人工费"
sheet14['B192'] = 1200000
sheet14['C192'] = 150000
sheet14['D192'] = 1000000
sheet14['E192'] = 1150000
sheet14['F192'] = 50000
sheet14['G192'] = "4%"
sheet14['H192'] = "李工"
sheet14['I192'] = "正常"
sheet14['A193'] = "排屋-材料费"
sheet14['B193'] = 8500000
sheet14['C193'] = 3000000
sheet14['D193'] = 5800000
sheet14['E193'] = 8800000
sheet14['F193'] = -300000
sheet14['G193'] = "-4%"
sheet14['H193'] = "王工"
sheet14['I193'] = "超支"
sheet14['J193'] = "预警"
sheet14['A194'] = "商业-人工费"
sheet14['B194'] = 1800000
sheet14['C194'] = 200000
sheet14['D194'] = 1550000
sheet14['E194'] = 1750000
sheet14['F194'] = 50000
sheet14['G194'] = "3%"
sheet14['H194'] = "李工"
sheet14['I194'] = "正常"
sheet14['A195'] = "商业-材料费"
sheet14['B195'] = 12000000
sheet14['C195'] = 4000000
sheet14['D195'] = 8200000
sheet14['E195'] = 12200000
sheet14['F195'] = -200000
sheet14['G195'] = "-2%"
sheet14['H195'] = "王工"
sheet14['I195'] = "超支"
sheet14['J195'] = "预警"
sheet14['A196'] = "措施费"
sheet14['B196'] = 500000
sheet14['C196'] = 100000
sheet14['D196'] = 450000
sheet14['E196'] = 550000
sheet14['F196'] = -50000
sheet14['G196'] = "-10%"
sheet14['H196'] = "张工"
sheet14['I196'] = "超支"
sheet14['J196'] = "预警"
sheet14['A197'] = "..."
sheet14['B197'] = "..."
sheet14['C197'] = "..."
sheet14['D197'] = "..."
sheet14['E197'] = "..."
sheet14['F197'] = "..."
sheet14['G197'] = "..."
sheet14['H197'] = "..."
sheet14['I197'] = "..."
sheet14['J197'] = "..."
sheet14['A198'] = "总计"
sheet14['B198'] = 82753467
sheet14['C198'] = 12000000
sheet14['D198'] = 71503467
sheet14['E198'] = 83503467
sheet14['F198'] = -750000
sheet14['G198'] = "-0.9%"
sheet14['H198'] = "项目经理"
sheet14['I198'] = "超支"
sheet14['J198'] = "重点关注"

# 设置字体和对齐
for row in range(191, 199):
    if row == 191:
        for col in range(1, 11):
            cell = sheet14[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 11):
            cell = sheet14[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 创建核心参数工作表
sheet15 = wb.create_sheet(title="核心参数")
sheet15['A202'] = "参数类别"
sheet15['B202'] = "参数名称"
sheet15['C202'] = "数值"
sheet15['A203'] = "项目总体"
sheet15['B203'] = "总建筑面积"
sheet15['C203'] = "-"
sheet15['A204'] = "项目总体"
sheet15['B204'] = "幕墙总面积"
sheet15['C204'] = "~ 48,000 ㎡"
sheet15['A205'] = "项目总体"
sheet15['B205'] = "合同总金额（暂估）"
sheet15['C205'] = "84,002,317 元"
sheet15['A206'] = "项目总体"
sheet15['B206'] = "目标利润率"
sheet15['C206'] = "5%"
sheet15['A207'] = "成本控制"
sheet15['B207'] = "人工费占总成本比例"
sheet15['C207'] = "15%"
sheet15['A208'] = "成本控制"
sheet15['B208'] = "材料费占总成本比例"
sheet15['C208'] = "72%"
sheet15['A209'] = "成本控制"
sheet15['B209'] = "措施及管理费占总成本比例"
sheet15['C209'] = "8%"
sheet15['A210'] = "成本控制"
sheet15['B210'] = "主材损耗率控制目标"
sheet15['C210'] = "2% (铝板/石材)"
sheet15['A211'] = "分包管理"
sheet15['B211'] = "劳务分包下浮目标"
sheet15['C211'] = "5% - 10%"
sheet15['A212'] = "分包管理"
sheet15['B212'] = "材料采购下浮目标"
sheet15['C212'] = "8% - 15%"
sheet15['A213'] = "财务指标"
sheet15['B213'] = "月度进度款支付比例"
sheet15['C213'] = "80%"
sheet15['A214'] = "财务指标"
sheet15['B214'] = "质保金预留比例"
sheet15['C214'] = "3%"
sheet15['A215'] = "财务指标"
sheet15['B215'] = "企业所得税率"
sheet15['C215'] = "25%"
sheet15['A216'] = "财务指标"
sheet15['B216'] = "增值税率"
sheet15['C216'] = "9%"

# 设置字体和对齐
for row in range(202, 217):
    if row == 202:
        for col in range(1, 4):
            cell = sheet15[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 4):
            cell = sheet15[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 创建执行进度计划工作表
sheet16 = wb.create_sheet(title="执行进度计划")
sheet16['A220'] = "序号"
sheet16['B220'] = "任务阶段"
sheet16['C220'] = "开始时间"
sheet16['D220'] = "结束时间"
sheet16['E220'] = "前置任务"
sheet16['F220'] = "状态"
sheet16['A221'] = "1"
sheet16['B221'] = "前期策划与招标"
sheet16['C221'] = "2024-01-01"
sheet16['D221'] = "2024-02-28"
sheet16['F221'] = "已完成"
sheet16['A222'] = "1.1"
sheet16['B222'] = "成本测算与预算编制"
sheet16['C222'] = "2024-01-01"
sheet16['D222'] = "2024-01-15"
sheet16['F222'] = "已完成"
sheet16['A223'] = "1.2"
sheet16['B223'] = "分包商与供应商招标"
sheet16['C223'] = "2024-01-16"
sheet16['D223'] = "2024-02-28"
sheet16['E223'] = "1.1"
sheet16['F223'] = "进行中"
sheet16['A224'] = "2"
sheet16['B224'] = "设计提料与采购"
sheet16['C224'] = "2024-02-01"
sheet16['D224'] = "2024-05-31"
sheet16['E224'] = "1.2"
sheet16['A225'] = "2.1"
sheet16['B225'] = "深化设计及图纸确认"
sheet16['C225'] = "2024-02-01"
sheet16['D225'] = "2024-03-31"
sheet16['F225'] = "进行中"
sheet16['A226'] = "2.2"
sheet16['B226'] = "石材、铝板、玻璃下单"
sheet16['C226'] = "2024-02-15"
sheet16['D226'] = "2024-03-15"
sheet16['E226'] = "2.1"
sheet16['A227'] = "2.3"
sheet16['B227'] = "材料加工生产"
sheet16['C227'] = "2024-03-16"
sheet16['D227'] = "2024-05-31"
sheet16['E227'] = "2.2"
sheet16['A228'] = "3"
sheet16['B228'] = "现场施工"
sheet16['C228'] = "2024-04-01"
sheet16['D228'] = "2024-12-31"
sheet16['E228'] = "2.3"
sheet16['A229'] = "3.1"
sheet16['B229'] = "商业/排屋龙骨安装"
sheet16['C229'] = "2024-04-01"
sheet16['D229'] = "2024-07-31"
sheet16['A230'] = "3.2"
sheet16['B230'] = "商业/排屋面材安装"
sheet16['C230'] = "2024-06-01"
sheet16['D230'] = "2024-10-31"
sheet16['E230'] = "3.1"
sheet16['A231'] = "3.3"
sheet16['B231'] = "塔楼幕墙安装"
sheet16['C231'] = "2024-07-01"
sheet16['D231'] = "2024-12-31"
sheet16['A232'] = "4"
sheet16['B232'] = "竣工验收与结算"
sheet16['C232'] = "2025-01-01"
sheet16['D232'] = "2025-04-30"
sheet16['E232'] = "3"
sheet16['A233'] = "4.1"
sheet16['B233'] = "竣工清理与验收"
sheet16['C233'] = "2025-01-01"
sheet16['D233'] = "2025-01-31"
sheet16['A234'] = "4.2"
sheet16['B234'] = "竣工图及结算资料编制"
sheet16['C234'] = "2025-02-01"
sheet16['D234'] = "2025-04-30"
sheet16['E234'] = "4.1"

# 设置字体和对齐
for row in range(220, 235):
    if row == 220:
        for col in range(1, 7):
            cell = sheet16[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 7):
            cell = sheet16[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 保存为 .xlsx 文件
output_file = "/Users/rs/AI/工程成本管理工具/中天建设_塔子山幕墙_成本管理工具/中天建设_塔子山幕墙_成本管理工具.xlsx"
wb.save(output_file)
print(f"Excel 文件已生成: {output_file}")

# 使用 xlsxwriter 添加宏功能
import xlsxwriter

# 创建一个新的工作簿，包含宏
workbook = xlsxwriter.Workbook("/Users/rs/AI/工程成本管理工具/中天建设_塔子山幕墙_成本管理工具/中天建设_塔子山幕墙_成本管理工具.xlsm", {'macros': True})

# 复制所有工作表和数据
for sheet_name in wb.sheetnames:
    source_sheet = wb[sheet_name]
    target_sheet = workbook.add_worksheet(sheet_name)
    
    # 复制数据
    row_num = 0
    for row in source_sheet.iter_rows(values_only=True):
        if row:
            target_sheet.write_row(row_num, 0, row)
            row_num += 1

# 添加宏
workbook.define_name('Auto_Open', '= macros!$A$1')

# 创建宏工作表
macros_sheet = workbook.add_worksheet('macros')
macros_sheet.hide()

# 添加 VBA 代码
vba_code = '''
Sub Auto_Open()
    MsgBox "欢迎使用中天建设塔子山幕墙项目成本管理工具！", vbInformation, "系统提示"
End Sub

Sub 刷新数据()
    ' 刷新所有数据
    Application.CalculateFull
    MsgBox "数据已刷新！", vbInformation, "系统提示"
End Sub

Sub 导出报表()
    ' 导出报表功能
    MsgBox "报表导出功能开发中...", vbInformation, "系统提示"
End Sub
'''

# 写入 VBA 代码
macros_sheet.write('A1', vba_code)

# 关闭工作簿
workbook.close()

print("Excel 文件已转换为 .xlsm 格式并添加宏功能")
