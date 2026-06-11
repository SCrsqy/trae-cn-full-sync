import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import os

wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

# 项目封面
sheet1 = wb.create_sheet(title="项目封面")
sheet1['A4'] = "中天建设集团有限公司"
sheet1['A5'] = "成华区迎晖路194号住宅、商业及配套设施项目"
sheet1['A6'] = "幕墙专业分包工程"
sheet1['A7'] = "项目成本管理策划书"
sheet1['A8'] = "编制部门： 项目部/成本合约部"
sheet1['A9'] = "编制日期： 2024年X月X日"
sheet1['A10'] = "版本号： V1.0"
sheet1['A11'] = "【内部机密 请勿外传】"

for row in range(4, 12):
    cell = sheet1[f'A{row}']
    cell.font = Font(size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')

# 项目概况
sheet2 = wb.create_sheet(title="项目概况")
sheet2['A15'] = "一、项目基本信息"
sheet2['A16'] = "项目名称"
sheet2['B16'] = "成华区迎晖路194号住宅、商业及配套设施项目幕墙工程"
sheet2['A17'] = "工程地点"
sheet2['B17'] = "成都市成华区迎晖路194号"
sheet2['A18'] = "建设单位"
sheet2['B18'] = "长住集团"
sheet2['A19'] = "总包单位"
sheet2['B19'] = "中天建设集团有限公司西南公司"
sheet2['A20'] = "二、工程范围与规模"
sheet2['A21'] = "工程范围"
sheet2['B21'] = "本项目包含排屋（11#楼）、1#2#商业及1#2#塔楼的幕墙工程。"
sheet2['A22'] = "主要幕墙形式"
sheet2['B22'] = "玻璃幕墙（座槽式、T型钢式、明框式）、石材幕墙、铝板幕墙、铝合金门窗、格栅、玻璃栏杆、采光顶等。"
sheet2['A23'] = "暂估幕墙面积"
sheet2['B23'] = "1. 排屋: 3,653.56 ㎡"
sheet2['B24'] = "2. 1#2#商业: 8,319.41 ㎡"
sheet2['B25'] = "3. 1#2#塔楼: 30,816.26 ㎡"
sheet2['A26'] = "三、合同与造价"
sheet2['A27'] = "合同形式"
sheet2['B27'] = "暂定工程量，固定综合单价"
sheet2['A28'] = "项目总预算"
sheet2['B28'] = "84,002,317 元 (暂估，含塔楼)"
sheet2['A29'] = "四、项目重难点"
sheet2['A30'] = "1. 材料复杂性"
sheet2['B30'] = "石材、铝板、玻璃种类繁多，且存在大量弧形、异形板，加工和安装难度高。"
sheet2['A31'] = "2. 图纸不完善"
sheet2['B31'] = "目前图纸不齐全，工程量存在较大不确定性，是主要风险源。"
sheet2['A32'] = "3. 工期紧张"
sheet2['B32'] = "展示区要求高，存在赶工和夜间施工情况，将增加措施及管理成本。"
sheet2['A33'] = "4. 质量控制"
sheet2['B33'] = "高端住宅项目，对幕墙的观感质量、防水性能等要求极高。"

for row in range(15, 34):
    if row in [15, 20, 26, 29]:
        cell = sheet2[f'A{row}']
        cell.font = Font(size=12, bold=True)
    else:
        cell = sheet2[f'A{row}']
        cell.font = Font(size=11)
    sheet2[f'B{row}'].alignment = Alignment(wrap_text=True)

# 成本利润基础数据
sheet3 = wb.create_sheet(title="成本利润基础数据")
sheet3['A37'] = "一、成本汇总（不含税）"
sheet3['A38'] = "序号"
sheet3['B38'] = "项目"
sheet3['C38'] = "金额（元）"
sheet3['D38'] = "占总成本比例"
sheet3['E38'] = "备注"
sheet3['A39'] = "1"
sheet3['B39'] = "排屋幕墙直接成本"
sheet3['C39'] = 3425614
sheet3['D39'] = "22%"
sheet3['E39'] = "基于工程量清单（11#楼）"
sheet3['A40'] = "2"
sheet3['B40'] = "商业幕墙直接成本"
sheet3['C40'] = 7934003
sheet3['D40'] = "51%"
sheet3['E40'] = "基于工程量清单（1#2#商业）"
sheet3['A41'] = "3"
sheet3['B41'] = "塔楼幕墙直接成本"
sheet3['C41'] = 25390235
sheet3['D41'] = "16%"
sheet3['E41'] = "基于面积及单方造价估算"
sheet3['A42'] = "4"
sheet3['B42'] = "措施费"
sheet3['C42'] = 525204
sheet3['D42'] = "3%"
sheet3['E42'] = "脚手架、吊篮、搬运等"
sheet3['A43'] = "5"
sheet3['B43'] = "管理费"
sheet3['C43'] = 4900000
sheet3['D43'] = "3%"
sheet3['E43'] = "管理人员工资、差旅、办公等"
sheet3['A44'] = "6"
sheet3['B44'] = "规费及税金"
sheet3['C44'] = 7500000
sheet3['D44'] = "5%"
sheet3['E44'] = "按9%税率估算"
sheet3['A45'] = "7"
sheet3['B45'] = "项目总成本"
sheet3['C45'] = 41582056
sheet3['D45'] = "100%"
sheet3['A47'] = "二、收入与利润测算"
sheet3['A48'] = "1"
sheet3['B48'] = "预计总收入"
sheet3['C48'] = 84002317
sheet3['E48'] = "参考中标价或与业主暂定价"
sheet3['A49'] = "2"
sheet3['B49'] = "预计总成本"
sheet3['C49'] = 41582056
sheet3['A50'] = "3"
sheet3['B50'] = "预计利润"
sheet3['C50'] = 42420261
sheet3['A51'] = "4"
sheet3['B51'] = "预计利润率"
sheet3['C51'] = "50.50%"
sheet3['E51'] = "注：塔楼成本按估算值，实际以分包价为准"

for row in range(37, 52):
    if row in [37, 47]:
        cell = sheet3[f'A{row}']
        cell.font = Font(size=12, bold=True)
    elif row == 38:
        for col in range(1, 6):
            cell = sheet3[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        cell = sheet3[f'A{row}']
        cell.font = Font(size=11)

# 成本明细分析
sheet4 = wb.create_sheet(title="成本明细分析")
sheet4['A54'] = "序号"
sheet4['B54'] = "区域"
sheet4['C54'] = "项目名称"
sheet4['D54'] = "单位"
sheet4['E54'] = "工程量"
sheet4['F54'] = "预算收入"
sheet4['G54'] = ""
sheet4['H54'] = "目标成本"
sheet4['I54'] = ""
sheet4['J54'] = ""
sheet4['K54'] = ""
sheet4['L54'] = ""
sheet4['M54'] = "盈亏分析"
sheet4['F55'] = "单价"
sheet4['G55'] = "合价"
sheet4['H55'] = "其中:主材费"
sheet4['I55'] = "人工费"
sheet4['J55'] = "辅材机械费"
sheet4['K55'] = "管理费"
sheet4['L55'] = "合价"
sheet4['M55'] = "盈亏"
sheet4['M56'] = "备注"

# 排屋数据
sheet4['A56'] = "1"
sheet4['B56'] = "排屋"
sheet4['C56'] = "80系列断桥隔热铝合金窗"
sheet4['D56'] = "m²"
sheet4['E56'] = 115.204
sheet4['F56'] = 754
sheet4['G56'] = 86828.23
sheet4['H56'] = 452
sheet4['I56'] = 121
sheet4['J56'] = 30
sheet4['K56'] = 7
sheet4['L56'] = 69934
sheet4['M56'] = 16894.23

sheet4['A57'] = "2"
sheet4['B57'] = "排屋"
sheet4['C57'] = "幕墙式铝合金窗"
sheet4['D57'] = "m²"
sheet4['E57'] = 515.11
sheet4['F57'] = 1259
sheet4['G57'] = 648486.36
sheet4['H57'] = 755
sheet4['I57'] = 202
sheet4['J57'] = 50
sheet4['K57'] = 12
sheet4['L57'] = 521952
sheet4['M57'] = 126534.36

sheet4['A58'] = "3"
sheet4['B58'] = "排屋"
sheet4['C58'] = "石材幕墙"
sheet4['D58'] = "m²"
sheet4['E58'] = 511.92
sheet4['F58'] = 1199
sheet4['G58'] = 613586.53
sheet4['H58'] = 719
sheet4['I58'] = 192
sheet4['J58'] = 48
sheet4['K58'] = 11
sheet4['L58'] = 494000
sheet4['M58'] = 119586.53

sheet4['A59'] = "4"
sheet4['B59'] = "排屋"
sheet4['C59'] = "铝板幕墙"
sheet4['D59'] = "m²"
sheet4['E59'] = 2217.79
sheet4['F59'] = 693
sheet4['G59'] = 1536058.52
sheet4['H59'] = 416
sheet4['I59'] = 111
sheet4['J59'] = 28
sheet4['K59'] = 7
sheet4['L59'] = 1236446
sheet4['M59'] = 299612.52

sheet4['A60'] = "5"
sheet4['B60'] = "排屋"
sheet4['C60'] = "铝合金格栅/百叶"
sheet4['D60'] = "m²"
sheet4['E60'] = 73.55
sheet4['F60'] = 676
sheet4['G60'] = 49734.82
sheet4['H60'] = 406
sheet4['I60'] = 108
sheet4['J60'] = 27
sheet4['K60'] = 6
sheet4['L60'] = 40037
sheet4['M60'] = 9697.82

sheet4['A61'] = "6"
sheet4['B61'] = "排屋"
sheet4['C61'] = "铝合金线条"
sheet4['D61'] = "m"
sheet4['E61'] = 867.93
sheet4['F61'] = 66
sheet4['G61'] = 57575.41
sheet4['H61'] = 40
sheet4['I61'] = 11
sheet4['J61'] = 3
sheet4['K61'] = 1
sheet4['L61'] = 46352
sheet4['M61'] = 11223.41

sheet4['A62'] = "7"
sheet4['B62'] = "排屋"
sheet4['C62'] = "玻璃采光顶"
sheet4['D62'] = "m²"
sheet4['E62'] = 99.04
sheet4['F62'] = 832
sheet4['G62'] = 82457.32
sheet4['H62'] = 499
sheet4['I62'] = 133
sheet4['J62'] = 33
sheet4['K62'] = 8
sheet4['L62'] = 66376
sheet4['M62'] = 16081.32

sheet4['A63'] = "8"
sheet4['B63'] = "排屋"
sheet4['C63'] = "玻璃栏杆"
sheet4['D63'] = "m"
sheet4['E63'] = 120.94
sheet4['F63'] = 852
sheet4['G63'] = 103078.32
sheet4['H63'] = 511
sheet4['I63'] = 136
sheet4['J63'] = 34
sheet4['K63'] = 8
sheet4['L63'] = 82989
sheet4['M63'] = 20089.32

sheet4['A64'] = "9"
sheet4['B64'] = "排屋"
sheet4['C64'] = "其他"
sheet4['D64'] = "项"
sheet4['E64'] = 1
sheet4['F64'] = 57930
sheet4['G64'] = 57930.38
sheet4['H64'] = 34758
sheet4['I64'] = 9269
sheet4['J64'] = 2317
sheet4['K64'] = 558
sheet4['L64'] = 46642
sheet4['M64'] = 11288.38

sheet4['A65'] = "10"
sheet4['B65'] = "排屋"
sheet4['C65'] = "措施费"
sheet4['D65'] = "项"
sheet4['E65'] = 1
sheet4['F65'] = 189878
sheet4['G65'] = 189878
sheet4['H65'] = 113927
sheet4['I65'] = 30407
sheet4['J65'] = 7602
sheet4['K65'] = 1832
sheet4['L65'] = 152856
sheet4['M65'] = 37022

sheet4['A66'] = ""
sheet4['B66'] = "排屋小计"
sheet4['G66'] = 3425613.87
sheet4['L66'] = 2757584
sheet4['M66'] = 668029.87

# 商业数据
sheet4['A67'] = "11"
sheet4['B67'] = "商业"
sheet4['C67'] = "玻璃幕墙"
sheet4['D67'] = "m²"
sheet4['E67'] = 1528.55
sheet4['F67'] = 1264
sheet4['G67'] = 1931614.01
sheet4['H67'] = 758
sheet4['I67'] = 202
sheet4['J67'] = 51
sheet4['K67'] = 12
sheet4['L67'] = 1555139
sheet4['M67'] = 376475.01

sheet4['A68'] = "12"
sheet4['B68'] = "商业"
sheet4['C68'] = "玻璃地弹门"
sheet4['D68'] = "m²"
sheet4['E68'] = 138.49
sheet4['F68'] = 1052
sheet4['G68'] = 145653.30
sheet4['H68'] = 631
sheet4['I68'] = 168
sheet4['J68'] = 42
sheet4['K68'] = 10
sheet4['L68'] = 117274
sheet4['M68'] = 28379.30

sheet4['A69'] = "13"
sheet4['B69'] = "商业"
sheet4['C69'] = "石材幕墙"
sheet4['D69'] = "m²"
sheet4['E69'] = 862.11
sheet4['F69'] = 1641
sheet4['G69'] = 1414940.22
sheet4['H69'] = 985
sheet4['I69'] = 263
sheet4['J69'] = 66
sheet4['K69'] = 16
sheet4['L69'] = 1139325
sheet4['M69'] = 275615.22

sheet4['A70'] = "14"
sheet4['B70'] = "商业"
sheet4['C70'] = "金属幕墙（铝板/网）"
sheet4['D70'] = "m²"
sheet4['E70'] = 4634.49
sheet4['F70'] = 619
sheet4['G70'] = 2867597.08
sheet4['H70'] = 371
sheet4['I70'] = 99
sheet4['J70'] = 25
sheet4['K70'] = 6
sheet4['L70'] = 2308510
sheet4['M70'] = 559087.08

sheet4['A71'] = "15"
sheet4['B71'] = "商业"
sheet4['C71'] = "金属网"
sheet4['D71'] = "m²"
sheet4['E71'] = 431.63
sheet4['F71'] = 578
sheet4['G71'] = 249586.32
sheet4['H71'] = 347
sheet4['I71'] = 93
sheet4['J71'] = 23
sheet4['K71'] = 6
sheet4['L71'] = 200964
sheet4['M71'] = 48622.32

sheet4['A72'] = "16"
sheet4['B72'] = "商业"
sheet4['C72'] = "铝合金格栅与百叶"
sheet4['D72'] = "m²"
sheet4['E72'] = 140.30
sheet4['F72'] = 681
sheet4['G72'] = 95581.07
sheet4['H72'] = 409
sheet4['I72'] = 109
sheet4['J72'] = 27
sheet4['K72'] = 7
sheet4['L72'] = 76952
sheet4['M72'] = 18629.07

sheet4['A73'] = "17"
sheet4['B73'] = "商业"
sheet4['C73'] = "铝合金线条"
sheet4['D73'] = "m"
sheet4['E73'] = 6155.83
sheet4['F73'] = 63
sheet4['G73'] = 385753.24
sheet4['H73'] = 38
sheet4['I73'] = 10
sheet4['J73'] = 3
sheet4['K73'] = 1
sheet4['L73'] = 310576
sheet4['M73'] = 75177.24

sheet4['A74'] = "18"
sheet4['B74'] = "商业"
sheet4['C74'] = "不锈钢"
sheet4['D74'] = "m"
sheet4['E74'] = 491.00
sheet4['F74'] = 475
sheet4['G74'] = 233424.78
sheet4['H74'] = 285
sheet4['I74'] = 76
sheet4['J74'] = 19
sheet4['K74'] = 5
sheet4['L74'] = 187940
sheet4['M74'] = 45484.78

sheet4['A75'] = "19"
sheet4['B75'] = "商业"
sheet4['C75'] = "玻璃栏杆"
sheet4['D75'] = "m"
sheet4['E75'] = 92.83
sheet4['F75'] = 874
sheet4['G75'] = 81172.77
sheet4['H75'] = 525
sheet4['I75'] = 140
sheet4['J75'] = 35
sheet4['K75'] = 8
sheet4['L75'] = 65366
sheet4['M75'] = 15806.77

sheet4['A76'] = "20"
sheet4['B76'] = "商业"
sheet4['C76'] = "其他"
sheet4['D76'] = "项"
sheet4['E76'] = 1
sheet4['F76'] = 193354
sheet4['G76'] = 193353.83
sheet4['H76'] = 116013
sheet4['I76'] = 30963
sheet4['J76'] = 7741
sheet4['K76'] = 1865
sheet4['L76'] = 155678
sheet4['M76'] = 37675.83

sheet4['A77'] = "21"
sheet4['B77'] = "商业"
sheet4['C77'] = "措施费"
sheet4['D77'] = "项"
sheet4['E77'] = 1
sheet4['F77'] = 335326
sheet4['G77'] = 335326.31
sheet4['H77'] = 201196
sheet4['I77'] = 53696
sheet4['J77'] = 13424
sheet4['K77'] = 3234
sheet4['L77'] = 269992
sheet4['M77'] = 65334.31

sheet4['A78'] = ""
sheet4['B78'] = "商业小计"
sheet4['G78'] = 7934002.93
sheet4['L78'] = 6388716
sheet4['M78'] = 1545286.93

# 塔楼数据
sheet4['A79'] = "22"
sheet4['B79'] = "塔楼"
sheet4['C79'] = "塔楼幕墙"
sheet4['D79'] = "m²"
sheet4['E79'] = 30816.26
sheet4['F79'] = 1018
sheet4['G79'] = 31370688
sheet4['L79'] = 25096550
sheet4['M79'] = 6274138
sheet4['H79'] = ""
sheet4['I79'] = ""
sheet4['J79'] = ""
sheet4['K79'] = ""
sheet4['M79'] = "单价按商业综合单价*80%估算"

sheet4['A80'] = ""
sheet4['B80'] = "直接费合计"
sheet4['G80'] = 81586262
sheet4['L80'] = 65555842
sheet4['M80'] = 16030420

for row in range(54, 81):
    if row in [54, 55]:
        for col in range(1, 14):
            cell = sheet4[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 14):
            cell = sheet4[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 利润与风险分析
sheet5 = wb.create_sheet(title="利润与风险分析")
sheet5['A71'] = "一、利润敏感性分析"
sheet5['A72'] = "变动因素"
sheet5['B72'] = "变动幅度"
sheet5['C72'] = "对总成本影响（元）"
sheet5['D72'] = "利润率影响"
sheet5['E72'] = "备注"
sheet5['A73'] = "主材价格上涨"
sheet5['B73'] = "+5%"
sheet5['C73'] = "+2,500,000"
sheet5['D73'] = "-2.98%"
sheet5['E73'] = "风险极高，铝材、玻璃价格波动大"
sheet5['A74'] = "人工成本上涨"
sheet5['B74'] = "+10%"
sheet5['C74'] = "+800,000"
sheet5['D74'] = "-0.95%"
sheet5['E74'] = "赶工期导致人工成本增加"
sheet5['A75'] = "工程量增加"
sheet5['B75'] = "+5%"
sheet5['C75'] = "+3,100,000"
sheet5['D75'] = "-3.69%"
sheet5['E75'] = "图纸不齐，变更风险巨大"
sheet5['A76'] = "材料损耗率控制"
sheet5['B76'] = "-1%"
sheet5['C76'] = "-500,000"
sheet5['D76'] = "+0.60%"
sheet5['E76'] = "通过精细化管理，可显著降本增效"
sheet5['A77'] = "二、主要风险识别与应对"
sheet5['A78'] = "风险类别"
sheet5['B78'] = "风险描述"
sheet5['C78'] = "可能性"
sheet5['D78'] = "影响程度"
sheet5['E78'] = "应对措施"
sheet5['A79'] = "工程量风险"
sheet5['B79'] = "图纸不完善，暂定工程量与实际差异大"
sheet5['C79'] = "高"
sheet5['D79'] = "高"
sheet5['E79'] = "1. 尽快与业主/设计确认图纸和工程量。\n2. 合同中明确变更索赔条款。\n3. 内部进行精细化翻样，提前预警。"
sheet5['A80'] = "材料风险"
sheet5['B80'] = "石材、铝板、弯弧玻璃等加工周期长、损耗大、价格波动"
sheet5['C80'] = "高"
sheet5['D80'] = "高"
sheet5['E80'] = "1. 提前锁定大宗材料价格。\n2. 优化排版，减少损耗。\n3. 派驻专人驻场监造，保证质量和进度。"
sheet5['A81'] = "工期风险"
sheet5['B81'] = "展示区工期紧张，需赶工"
sheet5['C81'] = "高"
sheet5['D81'] = "中"
sheet5['E81'] = "1. 制定详细可行的进度计划。\n2. 储备应急班组，应对高峰期。\n3. 赶工方案提前报批，保留索赔权利。"
sheet5['A82'] = "质量风险"
sheet5['B82'] = "弧形、异形板块安装精度难控制"
sheet5['C82'] = "中"
sheet5['D82'] = "高"
sheet5['E82'] = "1. 采用高精度测量放线仪器。\n2. 编制专项施工方案，样板引路。\n3. 加强过程验收。"

for row in range(71, 83):
    if row in [71, 77]:
        cell = sheet5[f'A{row}']
        cell.font = Font(size=12, bold=True)
    elif row in [72, 78]:
        for col in range(1, 6):
            cell = sheet5[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 6):
            cell = sheet5[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)
    sheet5[f'E{row}'].alignment = Alignment(wrap_text=True)

# 分包谈判建议
sheet6 = wb.create_sheet(title="分包谈判建议")
sheet6['A94'] = "谈判项"
sheet6['B94'] = "我方目标"
sheet6['C94'] = "分包商心理价位"
sheet6['D94'] = "谈判策略"
sheet6['E94'] = "负责人"
sheet6['A95'] = "劳务安装单价"
sheet6['B95'] = "在现有预算基础上下浮 5%-10%"
sheet6['C95'] = "预计可接受下浮 3%-5%"
sheet6['D95'] = "1. 以量换价：强调本项目体量大，系统多，可保证长期稳定用工。\n2. 难度加价反制：承认弧形、异形板块安装难度，但指出我方将提供完善的技术支持和测量放线，降低其管理成本。\n3. 付款条件：承诺进度款支付及时，换取价格优惠。"
sheet6['E95'] = "项目经理、成本经理"
sheet6['A96'] = "主材采购"
sheet6['B96'] = "争取 10%-15% 的下浮空间"
sheet6['C96'] = "不同材料下浮空间不同，石材5%，铝材8%，玻璃10%"
sheet6['D96'] = "1. 集中采购：整合排屋、商业、塔楼所有相同材料需求，统一招标。\n2. 引入竞争：每个品类至少邀请3家以上合格供应商报价。\n3. 战略合作：承诺后续项目优先合作权。"
sheet6['E96'] = "采购经理"
sheet6['A97'] = "措施费"
sheet6['B97'] = "打包总价，或按实结算部分争取 10% 下浮"
sheet6['C97'] = "5%左右"
sheet6['D97'] = "1. 对脚手架、吊篮等大型设备，利用集团长期合作商优势谈价。\n2. 对零星措施项，尽量包含在劳务单价中，减少单独计价。"
sheet6['E97'] = "生产经理、成本经理"
sheet6['A98'] = "付款条件"
sheet6['B98'] = "月进度80%，结算95%"
sheet6['C98'] = "月进度85%，结算95%"
sheet6['D98'] = "1. 以公司良好的付款信誉作为谈判筹码。\n2. 对于配合度高的分包商，可适当放宽付款比例。"
sheet6['E98'] = "项目经理、财务"

for row in range(94, 99):
    if row == 94:
        for col in range(1, 6):
            cell = sheet6[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 6):
            cell = sheet6[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)
    sheet6[f'D{row}'].alignment = Alignment(wrap_text=True)

# 刚挂石材专项预算
sheet8 = wb.create_sheet(title="刚挂石材专项预算")
sheet8['A133'] = "序号"
sheet8['B133'] = "项目名称"
sheet8['C133'] = "单位"
sheet8['D133'] = "工程量"
sheet8['E133'] = "主材费"
sheet8['F133'] = "劳务费"
sheet8['G133'] = "综合单价"
sheet8['A134'] = "1"
sheet8['B134'] = "25mm葡萄牙灰石材"
sheet8['C134'] = "m²"
sheet8['D134'] = 500
sheet8['E134'] = 680
sheet8['F134'] = 210
sheet8['G134'] = 890
sheet8['A135'] = "2"
sheet8['B135'] = "25mm葡萄牙米黄石材"
sheet8['C135'] = "m²"
sheet8['D135'] = 800
sheet8['E135'] = 380
sheet8['F135'] = 210
sheet8['G135'] = 590
sheet8['A136'] = "..."
sheet8['B136'] = "..."
sheet8['C136'] = "..."
sheet8['D136'] = "..."
sheet8['E136'] = "..."
sheet8['F136'] = "..."
sheet8['G136'] = "..."

for row in range(133, 137):
    if row == 133:
        for col in range(1, 8):
            cell = sheet8[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 8):
            cell = sheet8[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 玻璃幕墙专项预算
sheet9 = wb.create_sheet(title="玻璃幕墙专项预算")
sheet9['A140'] = "序号"
sheet9['B140'] = "项目名称"
sheet9['C140'] = "单位"
sheet9['D140'] = "工程量"
sheet9['E140'] = "主材费"
sheet9['F140'] = "劳务费"
sheet9['G140'] = "综合单价"
sheet9['A141'] = "1"
sheet9['B141'] = "座槽式幕墙 (12+1.9SGP+12夹胶)"
sheet9['C141'] = "m²"
sheet9['D141'] = 500
sheet9['E141'] = 1200
sheet9['F141'] = 190
sheet9['G141'] = 1390
sheet9['A142'] = "2"
sheet9['B142'] = "T型钢幕墙 (8+1.52PVB+8夹胶)"
sheet9['C142'] = "m²"
sheet9['D142'] = 800
sheet9['E142'] = 700
sheet9['F142'] = 180
sheet9['G142'] = 880
sheet9['A143'] = "..."
sheet9['B143'] = "..."
sheet9['C143'] = "..."
sheet9['D143'] = "..."
sheet9['E143'] = "..."
sheet9['F143'] = "..."
sheet9['G143'] = "..."

for row in range(140, 144):
    if row == 140:
        for col in range(1, 8):
            cell = sheet9[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 8):
            cell = sheet9[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 其他专项安装预算
sheet10 = wb.create_sheet(title="其他专项安装预算")
sheet10['A147'] = "序号"
sheet10['B147'] = "项目名称"
sheet10['C147'] = "单位"
sheet10['D147'] = "工程量"
sheet10['E147'] = "主材费"
sheet10['F147'] = "劳务费"
sheet10['G147'] = "综合单价"
sheet10['A148'] = "1"
sheet10['B148'] = "铝合金格栅"
sheet10['C148'] = "m²"
sheet10['D148'] = 200
sheet10['E148'] = 450
sheet10['F148'] = 150
sheet10['G148'] = 600
sheet10['A149'] = "2"
sheet10['B149'] = "玻璃栏杆"
sheet10['C149'] = "m"
sheet10['D149'] = 500
sheet10['E149'] = 400
sheet10['F149'] = 135
sheet10['G149'] = 535
sheet10['A150'] = "..."
sheet10['B150'] = "..."
sheet10['C150'] = "..."
sheet10['D150'] = "..."
sheet10['E150'] = "..."
sheet10['F150'] = "..."
sheet10['G150'] = "..."

for row in range(147, 151):
    if row == 147:
        for col in range(1, 8):
            cell = sheet10[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 8):
            cell = sheet10[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 材料清单明细
sheet11 = wb.create_sheet(title="材料清单明细")
sheet11['A156'] = "序号"
sheet11['B156'] = "类别"
sheet11['C156'] = "材料名称"
sheet11['D156'] = "规格型号"
sheet11['E156'] = "单位"
sheet11['F156'] = "数量"
sheet11['G156'] = "单价(含税)"
sheet11['H156'] = "合价"
sheet11['A157'] = "1"
sheet11['B157'] = "石材"
sheet11['C157'] = "25mm葡萄牙灰石材"
sheet11['E157'] = "m²"
sheet11['F157'] = 366.64
sheet11['G157'] = 705
sheet11['H157'] = 258481.20
sheet11['A158'] = "2"
sheet11['B158'] = "石材"
sheet11['C158'] = "25mm葡萄牙灰石材（圆弧）"
sheet11['E158'] = "m²"
sheet11['F158'] = 54.89
sheet11['G158'] = 1585
sheet11['H158'] = 86995.65
sheet11['A159'] = "3"
sheet11['B159'] = "石材"
sheet11['C159'] = "25mm葡萄牙米黄石材"
sheet11['E159'] = "m²"
sheet11['F159'] = 134.78
sheet11['G159'] = 405
sheet11['H159'] = 54585.90
sheet11['A160'] = "4"
sheet11['B160'] = "石材"
sheet11['C160'] = "25mm葡萄牙米黄石材（圆弧）"
sheet11['E160'] = "m²"
sheet11['F160'] = 26.96
sheet11['G160'] = 1125
sheet11['H160'] = 30330.00
sheet11['A161'] = "5"
sheet11['B161'] = "石材"
sheet11['C161'] = "35mm蓝眼睛石材（凹槽）"
sheet11['E161'] = "m²"
sheet11['F161'] = 123.69
sheet11['G161'] = 605
sheet11['H161'] = 74832.45
sheet11['A162'] = "..."
sheet11['B162'] = "..."
sheet11['C162'] = "..."
sheet11['D162'] = "..."
sheet11['E162'] = "..."
sheet11['F162'] = "..."
sheet11['G162'] = "..."
sheet11['H162'] = "..."
sheet11['A163'] = "100"
sheet11['B163'] = "铝型材"
sheet11['C163'] = "粉末喷涂铝型材T5"
sheet11['E163'] = "kg"
sheet11['F163'] = 20000
sheet11['G163'] = 27
sheet11['H163'] = 540000
sheet11['A164'] = "101"
sheet11['B164'] = "铝型材"
sheet11['C164'] = "阳极氧化铝型材T6"
sheet11['E164'] = "kg"
sheet11['F164'] = 5000
sheet11['G164'] = 30
sheet11['H164'] = 150000
sheet11['A165'] = "..."
sheet11['B165'] = "..."
sheet11['C165'] = "..."
sheet11['D165'] = "..."
sheet11['E165'] = "..."
sheet11['F165'] = "..."
sheet11['G165'] = "..."
sheet11['H165'] = "..."
sheet11['A166'] = "200"
sheet11['B166'] = "玻璃"
sheet11['C166'] = "12Low-E+12A+12 中空"
sheet11['E166'] = "m²"
sheet11['F166'] = 800
sheet11['G166'] = 299
sheet11['H166'] = 239200
sheet11['A167'] = "201"
sheet11['B167'] = "玻璃"
sheet11['C167'] = "15+2.28SGP+15 夹胶（超宽）"
sheet11['E167'] = "m²"
sheet11['F167'] = 100
sheet11['G167'] = 1695
sheet11['H167'] = 169500
sheet11['A168'] = "..."
sheet11['B168'] = "..."
sheet11['C168'] = "..."
sheet11['D168'] = "..."
sheet11['E168'] = "..."
sheet11['F168'] = "..."
sheet11['G168'] = "..."
sheet11['H168'] = "..."

for row in range(156, 169):
    if row == 156:
        for col in range(1, 9):
            cell = sheet11[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 9):
            cell = sheet11[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 分包下浮比例分析
sheet12 = wb.create_sheet(title="分包下浮比例分析")
sheet12['A169'] = "分包项目"
sheet12['B169'] = "我方预算总价"
sheet12['C169'] = "分包首次报价"
sheet12['D169'] = "谈判后目标价"
sheet12['E169'] = "目标下浮比例"
sheet12['F169'] = "责任人"
sheet12['A170'] = "排屋劳务"
sheet12['B170'] = 3200000
sheet12['C170'] = 3800000
sheet12['D170'] = 3500000
sheet12['E170'] = "7.8%"
sheet12['F170'] = "李经理"
sheet12['A171'] = "商业劳务"
sheet12['B171'] = 4500000
sheet12['C171'] = 5200000
sheet12['D171'] = 4800000
sheet12['E171'] = "7.7%"
sheet12['F171'] = "李经理"
sheet12['A172'] = "石材供应"
sheet12['B172'] = 2100000
sheet12['C172'] = 2300000
sheet12['D172'] = 2000000
sheet12['E172'] = "13.0%"
sheet12['F172'] = "王经理"
sheet12['A173'] = "铝材供应"
sheet12['B173'] = 3500000
sheet12['C173'] = 3800000
sheet12['D173'] = 3300000
sheet12['E173'] = "13.2%"
sheet12['F173'] = "王经理"
sheet12['A174'] = "..."
sheet12['B174'] = "..."
sheet12['C174'] = "..."
sheet12['D174'] = "..."
sheet12['E174'] = "..."
sheet12['F174'] = "..."

for row in range(169, 175):
    if row == 169:
        for col in range(1, 7):
            cell = sheet12[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 7):
            cell = sheet12[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 分包分析
sheet13 = wb.create_sheet(title="分包分析")
sheet13['A178'] = "分包商名称"
sheet13['B178'] = "承包范围"
sheet13['C178'] = "报价（元）"
sheet13['D178'] = "技术能力"
sheet13['E178'] = "配合度"
sheet13['F178'] = "综合评分"
sheet13['A179'] = "XX劳务公司"
sheet13['B179'] = "排屋劳务"
sheet13['C179'] = 3800000
sheet13['D179'] = "B+"
sheet13['E179'] = "A-"
sheet13['F179'] = 8.5
sheet13['A180'] = "YY劳务公司"
sheet13['B180'] = "商业劳务"
sheet13['C180'] = 5200000
sheet13['D180'] = "A-"
sheet13['E180'] = "B+"
sheet13['F180'] = 8.8
sheet13['A181'] = "ZZ石材公司"
sheet13['B181'] = "石材供应"
sheet13['C181'] = 2300000
sheet13['D181'] = "A"
sheet13['E181'] = "A"
sheet13['F181'] = 9.2
sheet13['A182'] = "AA铝材公司"
sheet13['B182'] = "铝材供应"
sheet13['C182'] = 3800000
sheet13['D182'] = "A-"
sheet13['E182'] = "A-"
sheet13['F182'] = 8.9
sheet13['A183'] = "..."
sheet13['B183'] = "..."
sheet13['C183'] = "..."
sheet13['D183'] = "..."
sheet13['E183'] = "..."
sheet13['F183'] = "..."
sheet13['A184'] = "结论与选择"
sheet13['A185'] = "经综合评估，建议选择："
sheet13['A186'] = "1. YY劳务公司作为商业劳务承包商，其技术能力更强，能更好应对复杂节点。"
sheet13['A187'] = "2. ZZ石材公司虽然价格不是最低，但其质量和配合度最佳，可有效降低损耗风险。"

for row in range(178, 188):
    if row == 178:
        for col in range(1, 7):
            cell = sheet13[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    elif row in [184, 185]:
        cell = sheet13[f'A{row}']
        cell.font = Font(size=11, bold=True)
    else:
        for col in range(1, 7):
            cell = sheet13[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 动态成本分析表
sheet14 = wb.create_sheet(title="中天建设塔子山幕墙项目动态成本分析表")
sheet14['A191'] = "成本科目"
sheet14['B191'] = "预算成本"
sheet14['C191'] = "已发生成本"
sheet14['D191'] = "待发生成本"
sheet14['E191'] = "动态总成本"
sheet14['F191'] = "偏差"
sheet14['G191'] = "偏差率"
sheet14['H191'] = "责任人"
sheet14['I191'] = "趋势"
sheet14['J191'] = "预警"
sheet14['A192'] = "排屋-人工费"
sheet14['B192'] = 3425614 * 0.15
sheet14['C192'] = 150000
sheet14['D192'] = 1000000
sheet14['E192'] = 1150000
sheet14['F192'] = 50000
sheet14['G192'] = "4%"
sheet14['H192'] = "李工"
sheet14['I192'] = "正常"
sheet14['A193'] = "排屋-材料费"
sheet14['B193'] = 3425614 * 0.72
sheet14['C193'] = 3000000
sheet14['D193'] = 5800000
sheet14['E193'] = 8800000
sheet14['F193'] = -300000
sheet14['G193'] = "-4%"
sheet14['H193'] = "王工"
sheet14['I193'] = "超支"
sheet14['J193'] = "预警"
sheet14['A194'] = "商业-人工费"
sheet14['B194'] = 7934003 * 0.15
sheet14['C194'] = 200000
sheet14['D194'] = 1550000
sheet14['E194'] = 1750000
sheet14['F194'] = 50000
sheet14['G194'] = "3%"
sheet14['H194'] = "李工"
sheet14['I194'] = "正常"
sheet14['A195'] = "商业-材料费"
sheet14['B195'] = 7934003 * 0.72
sheet14['C195'] = 4000000
sheet14['D195'] = 8200000
sheet14['E195'] = 12200000
sheet14['F195'] = -200000
sheet14['G195'] = "-2%"
sheet14['H195'] = "王工"
sheet14['I195'] = "超支"
sheet14['J195'] = "预警"
sheet14['A196'] = "措施费"
sheet14['B196'] = 525204
sheet14['C196'] = 100000
sheet14['D196'] = 450000
sheet14['E196'] = 550000
sheet14['F196'] = -50000
sheet14['G196'] = "-10%"
sheet14['H196'] = "张工"
sheet14['I196'] = "超支"
sheet14['J196'] = "预警"
sheet14['A197'] = "..."
sheet14['B197'] = "..."
sheet14['C197'] = "..."
sheet14['D197'] = "..."
sheet14['E197'] = "..."
sheet14['F197'] = "..."
sheet14['G197'] = "..."
sheet14['H197'] = "..."
sheet14['I197'] = "..."
sheet14['J197'] = "..."
sheet14['A198'] = "总计"
sheet14['B198'] = 41582056
sheet14['C198'] = 12000000
sheet14['D198'] = 71503467
sheet14['E198'] = 83503467
sheet14['F198'] = -750000
sheet14['G198'] = "-0.9%"
sheet14['H198'] = "项目经理"
sheet14['I198'] = "超支"
sheet14['J198'] = "重点关注"

for row in range(191, 199):
    if row == 191:
        for col in range(1, 11):
            cell = sheet14[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 11):
            cell = sheet14[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 核心参数
sheet15 = wb.create_sheet(title="核心参数")
sheet15['A202'] = "参数类别"
sheet15['B202'] = "参数名称"
sheet15['C202'] = "数值"
sheet15['A203'] = "项目总体"
sheet15['B203'] = "总建筑面积"
sheet15['C203'] = "-"
sheet15['A204'] = "项目总体"
sheet15['B204'] = "幕墙总面积"
sheet15['C204'] = "~ 42,789 ㎡ (排屋+商业)"
sheet15['A205'] = "项目总体"
sheet15['B205'] = "合同总金额（暂估）"
sheet15['C205'] = "84,002,317 元"
sheet15['A206'] = "项目总体"
sheet15['B206'] = "目标利润率"
sheet15['C206'] = "5%"
sheet15['A207'] = "成本控制"
sheet15['B207'] = "人工费占总成本比例"
sheet15['C207'] = "15%"
sheet15['A208'] = "成本控制"
sheet15['B208'] = "材料费占总成本比例"
sheet15['C208'] = "72%"
sheet15['A209'] = "成本控制"
sheet15['B209'] = "措施及管理费占总成本比例"
sheet15['C209'] = "8%"
sheet15['A210'] = "成本控制"
sheet15['B210'] = "主材损耗率控制目标"
sheet15['C210'] = "2% (铝板/石材)"
sheet15['A211'] = "分包管理"
sheet15['B211'] = "劳务分包下浮目标"
sheet15['C211'] = "5% - 10%"
sheet15['A212'] = "分包管理"
sheet15['B212'] = "材料采购下浮目标"
sheet15['C212'] = "8% - 15%"
sheet15['A213'] = "财务指标"
sheet15['B213'] = "月度进度款支付比例"
sheet15['C213'] = "80%"
sheet15['A214'] = "财务指标"
sheet15['B214'] = "质保金预留比例"
sheet15['C214'] = "3%"
sheet15['A215'] = "财务指标"
sheet15['B215'] = "企业所得税率"
sheet15['C215'] = "25%"
sheet15['A216'] = "财务指标"
sheet15['B216'] = "增值税率"
sheet15['C216'] = "9%"

for row in range(202, 217):
    if row == 202:
        for col in range(1, 4):
            cell = sheet15[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 4):
            cell = sheet15[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

# 执行进度计划
sheet16 = wb.create_sheet(title="执行进度计划")
sheet16['A220'] = "序号"
sheet16['B220'] = "任务阶段"
sheet16['C220'] = "开始时间"
sheet16['D220'] = "结束时间"
sheet16['E220'] = "前置任务"
sheet16['F220'] = "状态"
sheet16['A221'] = "1"
sheet16['B221'] = "前期策划与招标"
sheet16['C221'] = "2024-01-01"
sheet16['D221'] = "2024-02-28"
sheet16['F221'] = "已完成"
sheet16['A222'] = "1.1"
sheet16['B222'] = "成本测算与预算编制"
sheet16['C222'] = "2024-01-01"
sheet16['D222'] = "2024-01-15"
sheet16['F222'] = "已完成"
sheet16['A223'] = "1.2"
sheet16['B223'] = "分包商与供应商招标"
sheet16['C223'] = "2024-01-16"
sheet16['D223'] = "2024-02-28"
sheet16['E223'] = "1.1"
sheet16['F223'] = "进行中"
sheet16['A224'] = "2"
sheet16['B224'] = "设计提料与采购"
sheet16['C224'] = "2024-02-01"
sheet16['D224'] = "2024-05-31"
sheet16['E224'] = "1.2"
sheet16['A225'] = "2.1"
sheet16['B225'] = "深化设计及图纸确认"
sheet16['C225'] = "2024-02-01"
sheet16['D225'] = "2024-03-31"
sheet16['F225'] = "进行中"
sheet16['A226'] = "2.2"
sheet16['B226'] = "石材、铝板、玻璃下单"
sheet16['C226'] = "2024-02-15"
sheet16['D226'] = "2024-03-15"
sheet16['E226'] = "2.1"
sheet16['A227'] = "2.3"
sheet16['B227'] = "材料加工生产"
sheet16['C227'] = "2024-03-16"
sheet16['D227'] = "2024-05-31"
sheet16['E227'] = "2.2"
sheet16['A228'] = "3"
sheet16['B228'] = "现场施工"
sheet16['C228'] = "2024-04-01"
sheet16['D228'] = "2024-12-31"
sheet16['E228'] = "2.3"
sheet16['A229'] = "3.1"
sheet16['B229'] = "商业/排屋龙骨安装"
sheet16['C229'] = "2024-04-01"
sheet16['D229'] = "2024-07-31"
sheet16['A230'] = "3.2"
sheet16['B230'] = "商业/排屋面材安装"
sheet16['C230'] = "2024-06-01"
sheet16['D230'] = "2024-10-31"
sheet16['E230'] = "3.1"
sheet16['A231'] = "3.3"
sheet16['B231'] = "塔楼幕墙安装"
sheet16['C231'] = "2024-07-01"
sheet16['D231'] = "2024-12-31"
sheet16['A232'] = "4"
sheet16['B232'] = "竣工验收与结算"
sheet16['C232'] = "2025-01-01"
sheet16['D232'] = "2025-04-30"
sheet16['E232'] = "3"
sheet16['A233'] = "4.1"
sheet16['B233'] = "竣工清理与验收"
sheet16['C233'] = "2025-01-01"
sheet16['D233'] = "2025-01-31"
sheet16['A234'] = "4.2"
sheet16['B234'] = "竣工图及结算资料编制"
sheet16['C234'] = "2025-02-01"
sheet16['D234'] = "2025-04-30"
sheet16['E234'] = "4.1"

for row in range(220, 235):
    if row == 220:
        for col in range(1, 7):
            cell = sheet16[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    else:
        for col in range(1, 7):
            cell = sheet16[f'{get_column_letter(col)}{row}']
            cell.font = Font(size=11)

output_file = "/Users/rs/AI/工程成本管理工具/中天建设_塔子山幕墙_成本管理工具/中天建设_塔子山幕墙_成本管理工具_测算版.xlsx"
wb.save(output_file)
print(f"Excel 文件已生成: {output_file}")
