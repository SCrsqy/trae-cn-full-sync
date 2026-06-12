"""
统计分析预测报告生成器
自动生成包含统计分析、推算、预测数据的Excel报告
支持澳门(M)和香港(HK)两个数据源
"""
import sqlite3
import json
import time
import logging
from datetime import datetime
from typing import Dict, List, Optional, Any
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from time_weighted_analyzer import TimeWeightedAnalyzer, DrawRecord
from event_driven_accuracy import AccuracyCalculator
from data_sources import DataSourceManager, DATA_SOURCES

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 生肖映射
ZODIAC_MAPPING = {
    '鼠': [7, 19, 31, 43],
    '牛': [6, 18, 30, 42],
    '虎': [5, 17, 29, 41],
    '兔': [4, 16, 28, 40],
    '龙': [3, 15, 27, 39],
    '蛇': [2, 14, 26, 38],
    '马': [1, 13, 25, 37, 49],
    '羊': [12, 24, 36, 48],
    '猴': [11, 23, 35, 47],
    '鸡': [10, 22, 34, 46],
    '狗': [9, 21, 33, 45],
    '猪': [8, 20, 32, 44]
}

# 数字到生肖反向映射
NUMBER_TO_ZODIAC = {}
for zodiac, numbers in ZODIAC_MAPPING.items():
    for num in numbers:
        NUMBER_TO_ZODIAC[num] = zodiac


class ReportGenerator:
    """
    报告生成器
    生成包含统计分析、推算预测的Excel报告
    支持澳门(M)和香港(HK)两个数据源
    """

    def __init__(self, db_path: str, output_path: str, sources: List[str] = None):
        self.db_path = db_path
        self.output_path = output_path
        self.sources = sources or ['AM', 'HK']  # 默认两个数据源都生成
        self.analyzer = TimeWeightedAnalyzer(halflife_days=30)
        self.accuracy_calculator = AccuracyCalculator(ZODIAC_MAPPING)
        self.source_manager = DataSourceManager()

        # 样式定义
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')

        # 不同数据源的样式颜色
        self.source_colors = {
            'AM': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),  # 蓝色
            'HK': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),  # 绿色
        }

    def get_connection(self):
        """获取数据库连接"""
        return sqlite3.connect(self.db_path)

    def load_history_data(self, source_id: str = 'AM', limit: int = 200) -> List[DrawRecord]:
        """加载历史数据"""
        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute(f"""
                SELECT issue_number, draw_date, numbers, zodiacs
                FROM lottery_data_{source_id.lower()}
                ORDER BY issue_number DESC
                LIMIT ?
            """, (limit,))

            draws = []
            for row in cursor.fetchall():
                issue_number, draw_date, numbers_str, zodiacs_str = row

                # 解析JSON字符串
                try:
                    numbers = json.loads(numbers_str) if isinstance(numbers_str, str) else numbers_str
                    zodiacs = json.loads(zodiacs_str) if isinstance(zodiacs_str, str) else zodiacs_str
                except:
                    continue

                draws.append(DrawRecord(
                    issue_number=issue_number,
                    draw_date=datetime.fromisoformat(draw_date) if isinstance(draw_date, str) else draw_date,
                    numbers=numbers,
                    zodiacs=zodiacs
                ))

            # 按日期升序排列(用于分析)
            draws.reverse()
            return draws

        finally:
            conn.close()

    def calc_statistics(self, draws: List[DrawRecord]) -> Dict[str, Any]:
        """计算统计数据"""
        if not draws:
            return {}

        # 数字频率统计
        num_freq = self.analyzer.calc_number_frequency(draws)

        # 生肖频率统计
        zodiac_freq = self.analyzer.calc_zodiac_frequency(draws)

        # 三肖组合频率
        combo_freq = self.analyzer.calc_triple_combo_frequency(draws)
        top_combos = sorted(combo_freq.items(), key=lambda x: -x[1])[:10]

        return {
            'total_issues': len(draws),
            'num_frequency': num_freq,
            'zodiac_frequency': zodiac_freq,
            'top_combinations': top_combos,
            'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    def generate_prediction(self, stats: Dict) -> Dict[str, Any]:
        """生成预测数据"""
        if not stats:
            return {}

        # Top 6 生肖
        top6_zodiacs = [r.item for r in stats['zodiac_frequency'][:6]]

        # Top 12 数字
        top12_numbers = [r.item for r in stats['num_frequency'][:12]]

        # 4组三肖(不重叠)
        used_zodiacs = set()
        triple4_groups = []
        for combo, _ in stats['top_combinations']:
            if len(triple4_groups) >= 4:
                break
            if not set(combo) & used_zodiacs:
                triple4_groups.append(list(combo))
                used_zodiacs.update(combo)

        # 生成每组对应的数字
        triple4_with_numbers = []
        for group in triple4_groups:
            numbers = []
            for zodiac in group:
                numbers.extend(ZODIAC_MAPPING.get(zodiac, []))
            triple4_with_numbers.append({
                'group': group,
                'numbers': list(set(numbers))
            })

        return {
            'top6_zodiacs': top6_zodiacs,
            'top12_numbers': top12_numbers,
            'triple4_groups': triple4_with_numbers
        }

    def calc_prediction_accuracy(self, source_id: str = 'AM') -> Dict[str, Any]:
        """计算预测准确率"""
        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            # 获取最近的预测和实际数据
            cursor.execute(f"""
                SELECT p.issue_number, p.top6_zodiacs, p.triple4_groups, p.top12_numbers,
                       l.numbers as actual_numbers, l.zodiacs as actual_zodiacs,
                       p.accuracy_rate
                FROM predictions_{source_id.lower()} p
                LEFT JOIN lottery_data_{source_id.lower()} l ON p.issue_number = l.issue_number
                WHERE l.issue_number IS NOT NULL
                ORDER BY p.issue_number DESC
                LIMIT 30
            """, )

            rows = cursor.fetchall()

            if not rows:
                return {}

            accuracy_list = []
            for row in rows:
                issue, top6, triple4, top12, actual_nums, actual_zodiacs, acc = row
                accuracy_list.append({
                    'issue_number': issue,
                    'accuracy_rate': acc or 0,
                    'top6': json.loads(top6) if isinstance(top6, str) else top6,
                    'actual_zodiacs': json.loads(actual_zodiacs) if isinstance(actual_zodiacs, str) else actual_zodiacs
                })

            # 计算平均准确率
            avg_accuracy = sum(a['accuracy_rate'] for a in accuracy_list) / len(accuracy_list)
            max_accuracy = max(a['accuracy_rate'] for a in accuracy_list)
            min_accuracy = min(a['accuracy_rate'] for a in accuracy_list)

            return {
                'avg_accuracy': avg_accuracy,
                'max_accuracy': max_accuracy,
                'min_accuracy': min_accuracy,
                'recent_accuracies': accuracy_list[:10]
            }

        finally:
            conn.close()

    def write_to_excel(self, all_stats: Dict, all_predictions: Dict, all_accuracies: Dict):
        """写入Excel文件(支持多数据源)"""
        if openpyxl is None:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()

        # 删除默认sheet
        wb.remove(wb.active)

        # 为每个数据源创建Sheet
        for source in self.sources:
            source_name = "澳门" if source == "AM" else "香港"
            stats = all_stats.get(source, {})
            prediction = all_predictions.get(source, {})
            accuracy = all_accuracies.get(source, {})

            # 创建数据源Sheet
            self._create_source_sheet(wb, source, source_name, stats, prediction, accuracy)

        # Sheet: 历史数据(合并)
        self._write_history_sheet_all(wb)

        # Sheet: 数据源配置
        self._write_source_config_sheet(wb)

        # Sheet: 系统信息
        self._write_info_sheet_all(wb, all_stats, all_predictions, all_accuracies)

        # 保存文件
        wb.save(self.output_path)
        logger.info(f"报告已保存: {self.output_path}")

    def _create_source_sheet(self, wb, source: str, source_name: str, stats: Dict, prediction: Dict, accuracy: Dict):
        """为单个数据源创建Sheet"""
        ws = wb.create_sheet(f"{source_name}_{source}")

        row = 1
        # 标题
        ws.cell(row=row, column=1, value=f"{source_name}({source})统计分析预测报告").font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws.cell(row=row, column=1, value=f"统计期数: {stats.get('total_issues', 0)}")
        row += 1
        ws.cell(row=row, column=1, value=f"最后更新: {stats.get('last_update', '')}")
        row += 2

        # 数字频率统计
        ws.cell(row=row, column=1, value="数字频率统计").font = Font(bold=True)
        row += 1

        headers = ['排名', '数字', '出现次数', '简单频率', '加权频率']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.source_colors.get(source, self.header_fill)
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        for i, freq in enumerate(stats.get('num_frequency', [])[:20], 1):
            ws.cell(row=row, column=1, value=i).border = self.border
            ws.cell(row=row, column=2, value=freq.item).border = self.border
            ws.cell(row=row, column=3, value=freq.simple_count).border = self.border
            ws.cell(row=row, column=4, value=f"{freq.simple_frequency:.4f}").border = self.border
            ws.cell(row=row, column=5, value=f"{freq.weighted_frequency:.4f}").border = self.border
            row += 1

        row += 2

        # 生肖频率统计
        ws.cell(row=row, column=1, value="生肖频率统计").font = Font(bold=True)
        row += 1

        headers = ['排名', '生肖', '出现次数', '简单频率', '加权频率']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.source_colors.get(source, self.header_fill)
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        for i, freq in enumerate(stats.get('zodiac_frequency', []), 1):
            ws.cell(row=row, column=1, value=i).border = self.border
            ws.cell(row=row, column=2, value=freq.item).border = self.border
            ws.cell(row=row, column=3, value=freq.simple_count).border = self.border
            ws.cell(row=row, column=4, value=f"{freq.simple_frequency:.4f}").border = self.border
            ws.cell(row=row, column=5, value=f"{freq.weighted_frequency:.4f}").border = self.border
            row += 1

        row += 2

        # 预测推荐
        ws.cell(row=row, column=1, value="预测推荐").font = Font(bold=True)
        row += 1

        # 特肖6只
        ws.cell(row=row, column=1, value="特肖6只:").font = Font(bold=True)
        row += 1
        top6 = prediction.get('top6_zodiacs', [])
        for i, zodiac in enumerate(top6, 1):
            ws.cell(row=row, column=i, value=zodiac)
            ws.cell(row=row, column=i).alignment = self.center_align
            ws.cell(row=row, column=i).border = self.border
        row += 2

        # 4组三肖
        ws.cell(row=row, column=1, value="4组三肖:").font = Font(bold=True)
        row += 1
        headers = ['组别', '生肖', '对应数字']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.source_colors.get(source, self.header_fill)
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        for i, group_data in enumerate(prediction.get('triple4_groups', []), 1):
            ws.cell(row=row, column=1, value=f"第{i}组").border = self.border
            ws.cell(row=row, column=2, value='/'.join(group_data['group'])).border = self.border
            ws.cell(row=row, column=3, value=str(group_data['numbers'])).border = self.border
            row += 1

        row += 2

        # 热门12数字
        ws.cell(row=row, column=1, value="热门12数字:").font = Font(bold=True)
        row += 1
        top12 = prediction.get('top12_numbers', [])
        for i, num in enumerate(top12, 1):
            col = (i - 1) % 6 + 1
            r = row + (i - 1) // 6
            cell = ws.cell(row=r, column=col, value=num)
            cell.alignment = self.center_align
            cell.border = self.border

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

    def _write_statistics_sheet(self, wb, stats: Dict):
        """写入统计分析Sheet"""
        ws = wb.active
        ws.title = "统计分析"

        # 标题
        ws['A1'] = "统计分析报告"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')

        ws['A2'] = f"统计期数: {stats.get('total_issues', 0)}"
        ws['A3'] = f"最后更新: {stats.get('last_update', '')}"

        # 数字频率
        row = 5
        ws.cell(row=row, column=1, value="数字频率统计").font = Font(bold=True)
        row += 1

        headers = ['排名', '数字', '出现次数', '简单频率', '加权频率']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        row += 1
        for i, freq in enumerate(stats.get('num_frequency', [])[:20], 1):
            ws.cell(row=row, column=1, value=i).border = self.border
            ws.cell(row=row, column=2, value=freq.item).border = self.border
            ws.cell(row=row, column=3, value=freq.simple_count).border = self.border
            ws.cell(row=row, column=4, value=f"{freq.simple_frequency:.4f}").border = self.border
            ws.cell(row=row, column=5, value=f"{freq.weighted_frequency:.4f}").border = self.border
            row += 1

        # 生肖频率
        row += 2
        ws.cell(row=row, column=1, value="生肖频率统计").font = Font(bold=True)
        row += 1

        headers = ['排名', '生肖', '出现次数', '简单频率', '加权频率']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        row += 1
        for i, freq in enumerate(stats.get('zodiac_frequency', []), 1):
            ws.cell(row=row, column=1, value=i).border = self.border
            ws.cell(row=row, column=2, value=freq.item).border = self.border
            ws.cell(row=row, column=3, value=freq.simple_count).border = self.border
            ws.cell(row=row, column=4, value=f"{freq.simple_frequency:.4f}").border = self.border
            ws.cell(row=row, column=5, value=f"{freq.weighted_frequency:.4f}").border = self.border
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

    def _write_prediction_sheet(self, wb, prediction: Dict):
        """写入预测推荐Sheet"""
        ws = wb.create_sheet("预测推荐")

        # 标题
        ws['A1'] = "预测推荐数据"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # 特肖6只
        row = 4
        ws.cell(row=row, column=1, value="特肖6只 (按准确率排序)").font = Font(bold=True)
        row += 1

        top6 = prediction.get('top6_zodiacs', [])
        for i, zodiac in enumerate(top6, 1):
            ws.cell(row=row, column=i, value=zodiac)
            ws.cell(row=row, column=i).alignment = self.center_align
            ws.cell(row=row, column=i).border = self.border

        # 4组三肖
        row += 2
        ws.cell(row=row, column=1, value="4组三肖推荐").font = Font(bold=True)
        row += 1

        headers = ['组别', '生肖', '对应数字']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        row += 1
        for i, group_data in enumerate(prediction.get('triple4_groups', []), 1):
            ws.cell(row=row, column=1, value=f"第{i}组").border = self.border
            ws.cell(row=row, column=2, value='/'.join(group_data['group'])).border = self.border
            ws.cell(row=row, column=3, value=str(group_data['numbers'])).border = self.border
            row += 1

        # 热门12数字
        row += 2
        ws.cell(row=row, column=1, value="热门12个数字").font = Font(bold=True)
        row += 1

        top12 = prediction.get('top12_numbers', [])
        for i, num in enumerate(top12, 1):
            col = (i - 1) % 6 + 1
            r = row + (i - 1) // 6
            cell = ws.cell(row=r, column=col, value=num)
            cell.alignment = self.center_align
            cell.border = self.border

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15

    def _write_accuracy_sheet(self, wb, accuracy: Dict):
        """写入准确率统计Sheet"""
        ws = wb.create_sheet("准确率统计")

        # 标题
        ws['A1'] = "预测准确率统计"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        if not accuracy:
            ws['A3'] = "暂无数据"
            return

        # 汇总统计
        ws['A3'] = "汇总统计"
        ws['A3'].font = Font(bold=True)

        ws['A4'] = "平均准确率:"
        ws['B4'] = f"{accuracy.get('avg_accuracy', 0):.2f}%"
        ws['A5'] = "最高准确率:"
        ws['B5'] = f"{accuracy.get('max_accuracy', 0):.2f}%"
        ws['A6'] = "最低准确率:"
        ws['B6'] = f"{accuracy.get('min_accuracy', 0):.2f}%"

        # 近期准确率
        row = 8
        ws.cell(row=row, column=1, value="近期10期准确率").font = Font(bold=True)
        row += 1

        headers = ['期号', '准确率', '预测生肖', '实际生肖']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        row += 1
        for item in accuracy.get('recent_accuracies', []):
            ws.cell(row=row, column=1, value=item['issue_number']).border = self.border
            ws.cell(row=row, column=2, value=f"{item['accuracy_rate']:.2f}%").border = self.border
            ws.cell(row=row, column=3, value=str(item.get('top6', []))).border = self.border
            ws.cell(row=row, column=4, value=str(item.get('actual_zodiacs', []))).border = self.border
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30

    def _write_history_sheet(self, wb):
        """写入历史数据Sheet"""
        ws = wb.create_sheet("历史数据")

        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT issue_number, draw_date, numbers, zodiacs
                FROM lottery_data_am
                ORDER BY issue_number DESC
                LIMIT 100
            """)

            row = 1
            headers = ['期号', '开奖日期', '开奖号码', '对应生肖']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align
                cell.border = self.border

            row += 1
            for row_data in cursor.fetchall():
                issue, date, numbers, zodiacs = row_data
                try:
                    numbers_list = json.loads(numbers) if isinstance(numbers, str) else numbers
                    zodiacs_list = json.loads(zodiacs) if isinstance(zodiacs, str) else zodiacs
                except:
                    continue

                ws.cell(row=row, column=1, value=issue).border = self.border
                ws.cell(row=row, column=2, value=date).border = self.border
                ws.cell(row=row, column=3, value=str(numbers_list)).border = self.border
                ws.cell(row=row, column=4, value=str(zodiacs_list)).border = self.border
                row += 1

        finally:
            conn.close()

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

    def _write_history_sheet_all(self, wb):
        """写入所有数据源的历史数据"""
        ws = wb.create_sheet("历史数据")

        row = 1
        headers = ['数据源', '期号', '开奖日期', '开奖号码', '对应生肖']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        conn = self.get_connection()
        cursor = conn.cursor()

        try:
            for source in self.sources:
                cursor.execute(f"""
                    SELECT issue_number, draw_date, numbers, zodiacs
                    FROM lottery_data_{source.lower()}
                    ORDER BY issue_number DESC
                    LIMIT 50
                """)

                source_name = "澳门" if source == "AM" else "香港"

                for row_data in cursor.fetchall():
                    issue, date, numbers, zodiacs = row_data
                    try:
                        numbers_list = json.loads(numbers) if isinstance(numbers, str) else numbers
                        zodiacs_list = json.loads(zodiacs) if isinstance(zodiacs, str) else zodiacs
                    except:
                        continue

                    ws.cell(row=row, column=1, value=source_name).border = self.border
                    ws.cell(row=row, column=2, value=issue).border = self.border
                    ws.cell(row=row, column=3, value=date).border = self.border
                    ws.cell(row=row, column=4, value=str(numbers_list)).border = self.border
                    ws.cell(row=row, column=5, value=str(zodiacs_list)).border = self.border
                    row += 1

        finally:
            conn.close()

        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25

    def _write_source_config_sheet(self, wb):
        """写入数据源配置信息"""
        ws = wb.create_sheet("数据源配置")

        ws['A1'] = "数据源配置"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')

        row = 3
        headers = ['地区', '代码', 'URL', '状态']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        for source in DATA_SOURCES:
            ws.cell(row=row, column=1, value=source.name).border = self.border
            ws.cell(row=row, column=2, value=source.code).border = self.border
            ws.cell(row=row, column=3, value=source.url).border = self.border
            ws.cell(row=row, column=4, value="启用" if source.enabled else "禁用").border = self.border
            row += 1

        row += 2
        ws.cell(row=row, column=1, value="说明:").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="• 澳门(M): https://2026kj.zkclhb.com:2025/am.html#toubu13")
        row += 1
        ws.cell(row=row, column=1, value="• 香港(K): https://2026kj.zkclhb.com:2025/hk.html#toubu13")
        row += 1
        ws.cell(row=row, column=1, value="• 报告自动包含两个数据源的统计分析、预测推荐和准确率统计")

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10

    def _write_info_sheet_all(self, wb, all_stats: Dict, all_predictions: Dict, all_accuracies: Dict):
        """写入系统信息Sheet"""
        ws = wb.create_sheet("系统信息")

        ws['A1'] = "系统信息"
        ws['A1'].font = self.title_font

        ws['A3'] = "生成时间:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        ws['A4'] = "数据来源:"
        ws['B4'] = ", ".join([f"{'澳门' if s == 'AM' else '香港'}({s})" for s in self.sources])

        total_issues = sum(stats.get('total_issues', 0) for stats in all_stats.values())
        ws['A5'] = "总统计期数:"
        ws['B5'] = total_issues

        ws['A6'] = "时间加权半衰期:"
        ws['B6'] = "30天"

        ws['A7'] = "预测算法:"
        ws['B7'] = "时间加权频率统计"

        ws['A8'] = "准确率计算:"
        ws['B8'] = "特肖×0.6 + 三肖×0.3 + 数字×0.1"

        ws['A10'] = "生肖与数字映射"
        ws['A10'].font = Font(bold=True)

        row = 11
        for zodiac, numbers in ZODIAC_MAPPING.items():
            ws.cell(row=row, column=1, value=zodiac)
            ws.cell(row=row, column=2, value=str(numbers))
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40

    def generate_report(self) -> bool:
        """生成完整报告"""
        try:
            logger.info("开始生成报告...")
            logger.info(f"数据源: {self.sources}")

            all_stats = {}
            all_predictions = {}
            all_accuracies = {}

            # 为每个数据源生成数据
            for source in self.sources:
                source_name = "澳门" if source == "AM" else "香港"
                logger.info(f"处理 {source_name} ({source}) 数据...")

                # 1. 加载历史数据
                draws = self.load_history_data(source)
                if not draws:
                    logger.warning(f"{source_name} 没有历史数据，跳过")
                    continue

                # 2. 计算统计数据
                stats = self.calc_statistics(draws)
                stats['source_name'] = source_name
                all_stats[source] = stats

                # 3. 生成预测
                prediction = self.generate_prediction(stats)
                all_predictions[source] = prediction

                # 4. 计算准确率
                accuracy = self.calc_prediction_accuracy(source)
                all_accuracies[source] = accuracy

            if not all_stats:
                logger.warning("没有任何数据源有数据")
                return False

            # 5. 写入Excel
            logger.info("写入Excel...")
            self.write_to_excel(all_stats, all_predictions, all_accuracies)

            logger.info("报告生成完成!")
            return True

        except Exception as e:
            logger.error(f"报告生成失败: {e}")
            import traceback
            traceback.print_exc()
            return False


class RealtimeReportUpdater:
    """
    实时报告更新器
    定期更新Excel报告
    """

    def __init__(self, db_path: str, output_path: str, interval: int = 60):
        self.generator = ReportGenerator(db_path, output_path)
        self.interval = interval
        self.running = False

    def start(self):
        """启动实时更新"""
        self.running = True
        logger.info(f"启动实时报告更新器 (间隔: {self.interval}秒)")

        while self.running:
            try:
                self.generator.generate_report()
            except Exception as e:
                logger.error(f"更新失败: {e}")

            # 等待下一次更新
            for _ in range(self.interval):
                if not self.running:
                    break
                time.sleep(1)

    def stop(self):
        """停止更新"""
        self.running = False
        logger.info("停止实时报告更新器")


# 使用示例
if __name__ == "__main__":
    db_path = '/Users/rs/AI/分析预测推荐/lottery.db'
    output_path = '/Users/rs/AI/分析预测推荐/统计分析预测报告.xlsx'

    # 生成单次报告
    generator = ReportGenerator(db_path, output_path)
    success = generator.generate_report()

    if success:
        print(f"报告已生成: {output_path}")
    else:
        print("报告生成失败")

    # 或者启动实时更新(每60秒)
    # updater = RealtimeReportUpdater(db_path, output_path, interval=60)
    # updater.start()