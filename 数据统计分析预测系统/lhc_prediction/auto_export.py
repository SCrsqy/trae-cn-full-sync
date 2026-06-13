import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from config import DATABASE_PATH

def get_latest_draw_records(region, limit=1):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT region, issue, draw_date, numbers, special_number, zodiac
        FROM draw_results
        WHERE region = ?
        ORDER BY issue DESC
        LIMIT ?
    ''', (region, limit))
    results = cursor.fetchall()
    conn.close()
    
    records = []
    for row in results:
        records.append({
            '地区': '澳门' if row[0] == 'am' else '香港',
            '期号': row[1],
            '开奖日期': row[2],
            '平码': row[3],
            '特码': row[4],
            '特码生肖': row[5]
        })
    return records

def get_latest_prediction(region, limit=1):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT region, issue, pred_type, content, actual_special, is_correct, accuracy
        FROM predictions
        WHERE region = ? AND actual_special IS NOT NULL
        ORDER BY issue DESC
        LIMIT ?
    ''', (region, limit))
    results = cursor.fetchall()
    conn.close()
    
    records = []
    for row in results:
        records.append({
            '地区': '澳门' if row[0] == 'am' else '香港',
            '期号': row[1],
            '预测类型': {'special_6': '特码6只', 'triple_zodiac': '三肖组合', 'hot_12': '热门12码', 'special_6_zodiac': '6肖特码', 'consecutive_3_zodiac': '3连肖'}.get(row[2], row[2]),
            '预测内容': row[3],
            '实际特码': row[4],
            '是否命中': '是' if row[5] else '否',
            '准确率': f'{row[6]:.2f}%' if row[6] else '--'
        })
    return records

def get_next_prediction(region):
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT region, issue, pred_type, content
        FROM predictions
        WHERE region = ? AND actual_special IS NULL
        ORDER BY issue DESC
        LIMIT 1
    ''', (region,))
    results = cursor.fetchall()
    conn.close()
    
    predictions = {}
    for row in results:
        if row[1] not in predictions:
            predictions[row[1]] = {'issue': row[1]}
        predictions[row[1]][row[2]] = row[3]
    return predictions

def export_daily_report():
    export_dir = os.path.join(os.path.dirname(DATABASE_PATH), 'exports')
    os.makedirs(export_dir, exist_ok=True)
    
    today = datetime.now().strftime('%Y%m%d')
    filename = os.path.join(export_dir, f'lhc_daily_report_{today}.xlsx')
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for region, region_name in [('am', '澳门'), ('hk', '香港')]:
            latest_draw = get_latest_draw_records(region)
            if latest_draw:
                df_draw = pd.DataFrame(latest_draw)
                df_draw.to_excel(writer, sheet_name=f'{region_name}开奖记录', index=False)
            
            latest_pred = get_latest_prediction(region)
            if latest_pred:
                df_pred = pd.DataFrame(latest_pred)
                df_pred.to_excel(writer, sheet_name=f'{region_name}预测记录', index=False)
            
            next_pred = get_next_prediction(region)
            if next_pred:
                rows = []
                for issue, pred_data in next_pred.items():
                    special_6 = pred_data.get('special_6', 'N/A')
                    triple_zodiac = pred_data.get('triple_zodiac', 'N/A')
                    hot_12 = pred_data.get('hot_12', 'N/A')
                    rows.append({
                        '地区': region_name,
                        '期号': issue,
                        '特码6只预测': special_6,
                        '三肖组合预测': triple_zodiac,
                        '热门12码预测': hot_12
                    })
                df_next = pd.DataFrame(rows)
                df_next.to_excel(writer, sheet_name=f'{region_name}下期预测', index=False)
    
    return filename

def update_excel_file():
    export_dir = os.path.join(os.path.dirname(DATABASE_PATH), 'exports')
    os.makedirs(export_dir, exist_ok=True)
    
    filename = os.path.join(export_dir, 'lhc_latest_report.xlsx')
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for region, region_name in [('am', '澳门'), ('hk', '香港')]:
            latest_draw = get_latest_draw_records(region, 10)
            if latest_draw:
                df_draw = pd.DataFrame(latest_draw)
                df_draw.to_excel(writer, sheet_name=f'{region_name}开奖记录', index=False)
                _apply_styles(writer.sheets[f'{region_name}开奖记录'])
            
            latest_pred = get_latest_prediction(region, 10)
            if latest_pred:
                df_pred = pd.DataFrame(latest_pred)
                df_pred.to_excel(writer, sheet_name=f'{region_name}预测记录', index=False)
                _apply_styles(writer.sheets[f'{region_name}预测记录'])
            
            next_pred = get_next_prediction(region)
            if next_pred:
                rows = []
                for issue, pred_data in next_pred.items():
                    rows.append({
                        '地区': region_name,
                        '期号': issue,
                        '特码6只预测': pred_data.get('special_6', 'N/A'),
                        '三肖组合预测': pred_data.get('triple_zodiac', 'N/A'),
                        '热门12码预测': pred_data.get('hot_12', 'N/A')
                    })
                df_next = pd.DataFrame(rows)
                df_next.to_excel(writer, sheet_name=f'{region_name}下期预测', index=False)
                _apply_styles(writer.sheets[f'{region_name}下期预测'])
    
    return filename

def _apply_styles(ws):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

if __name__ == '__main__':
    filename = export_daily_report()
    print(f'Excel报告已生成: {filename}')